import os
import struct
import codecs
import traceback
import openpyxl
from tkinter import filedialog as fd
from tkinter import messagebox as mb

import program.textSetting as textSetting


class ExcelWidget:
    def __init__(self, decryptFile, reloadFunc):
        self.decryptFile = decryptFile
        self.reloadFunc = reloadFunc

    def extract(self):
        filename = self.decryptFile.filename + ".xlsx"
        file_path = fd.asksaveasfilename(initialfile=filename, defaultextension="xlsx", filetypes=[("railData", "*.xlsx")])
        if not file_path:
            return
        wb = openpyxl.Workbook()

        # シート初期化
        defSheetNameList = wb.sheetnames
        for sheetName in defSheetNameList:
            wb.remove(wb[sheetName])

        # TabList
        tabList = textSetting.textList["railEditor"]["railComboValue"]
        if self.decryptFile.game == "LS":
            tabList = textSetting.textList["railEditor"]["railLsComboValue"]
        for index, tabName in enumerate(tabList):
            wb.create_sheet(index=index, title=tabName)
            try:
                self.extractRailDataInfo(index, wb[tabName])
            except Exception:
                mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
                return

        try:
            wb.save(file_path)
            mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I113"])
        except PermissionError:
            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E94"])

    def extractRailDataInfo(self, sheetIndex, ws):
        row = 1
        # BGM、配置情報
        if sheetIndex == 0:
            # ver
            ws.cell(row, 1).value = self.decryptFile.ver
            row += 2

            # BGM
            ws.cell(row, 1).value = "BGM"
            if self.decryptFile.game in ["BS", "CS", "RS"]:
                ws.cell(row, 2).value = self.decryptFile.musicCnt
            row += 1
            if self.decryptFile.game in ["LS", "BS"]:
                for musicInfo in self.decryptFile.musicList:
                    for idx, music in enumerate(musicInfo):
                        ws.cell(row, 1 + idx).value = music
                    row += 1
            row += 1

            # 車両の初期レール位置
            ws.cell(row, 1).value = "RailPos"
            ws.cell(row, 2).value = self.decryptFile.trainCnt
            row += 1
            for trainInfo in self.decryptFile.trainList:
                for idx, train in enumerate(trainInfo):
                    ws.cell(row, 1 + idx).value = train
                row += 1
            row += 1

            if self.decryptFile.game in ["BS", "CS", "RS"]:
                # ダミー位置？
                ws.cell(row, 1).value = "RailPos2"
                row += 1
                for trainInfo2 in self.decryptFile.trainList2:
                    for idx, train2 in enumerate(trainInfo2):
                        ws.cell(row, 1 + idx).value = train2
                    row += 1
                row += 1

                # 試運転、二人バトルの初期レール位置
                ws.cell(row, 1).value = "FreeRunOrVSPos"
                row += 1
                for trainInfo3 in self.decryptFile.trainList3:
                    for idx, train3 in enumerate(trainInfo3):
                        ws.cell(row, 1 + idx).value = train3
                    row += 1
                row += 1

                # 駅表示を始める番号
                ws.cell(row, 1).value = "stationNo"
                ws.cell(row, 2).value = self.decryptFile.stationNo
                row += 2

                if self.decryptFile.game == "BS":
                    ws.cell(row, 1).value = "RailPos4"
                    row += 1
                    for trainInfo4 in self.decryptFile.trainList4:
                        for idx, train4 in enumerate(trainInfo4):
                            ws.cell(row, 1 + idx).value = train4
                        row += 1
                    row += 1

                    ws.cell(row, 1).value = "stationNo2"
                    ws.cell(row, 2).value = self.decryptFile.stationNo2
                    row += 2

            # レール名
            if self.decryptFile.game in ["LS", "BS"]:
                ws.cell(row, 1).value = "railName"
                ws.cell(row, 2).value = self.decryptFile.railStationName
        # 要素１
        elif sheetIndex == 1:
            # else1
            if self.decryptFile.game == "LS":
                ws.cell(row, 1).value = "else1"
                ws.cell(row, 2).value = len(self.decryptFile.else1List)
                row += 1

                for idx, else1 in enumerate(self.decryptFile.else1List):
                    ws.cell(row, 1 + idx).value = else1
                row += 2
            else:
                ws.cell(row, 1).value = "else1-1"
                ws.cell(row, 2).value = self.decryptFile.else1List[0]
                row += 2

                ws.cell(row, 1).value = "else1-2"
                ws.cell(row, 2).value = len(self.decryptFile.else1List[1:])
                row += 1

                for else1Info in self.decryptFile.else1List[1:]:
                    for idx, else1 in enumerate(else1Info):
                        ws.cell(row, 1 + idx).value = else1
                    row += 1
                row += 1

                # light情報
                ws.cell(row, 1).value = "light"
                ws.cell(row, 2).value = len(self.decryptFile.lightList)
                row += 1

                for lightName in self.decryptFile.lightList:
                    ws.cell(row, 1).value = lightName
                    row += 1
                row += 1

                if self.decryptFile.game in ["CS", "RS"]:
                    # 駅名標画像情報
                    ws.cell(row, 1).value = "StageRes"
                    ws.cell(row, 2).value = len(self.decryptFile.pngList)
                    row += 1

                    for pngName in self.decryptFile.pngList:
                        ws.cell(row, 1).value = pngName
                        row += 1
                    row += 1

                    # 駅名標AMB情報
                    ws.cell(row, 1).value = "SetTexInfo"
                    ws.cell(row, 2).value = len(self.decryptFile.stationList)
                    row += 1

                    for stationAmbInfo in self.decryptFile.stationList:
                        for idx, stationAmb in enumerate(stationAmbInfo):
                            ws.cell(row, 1 + idx).value = stationAmb
                        row += 1
                    row += 1

                # base bin
                ws.cell(row, 1).value = "baseBin"
                ws.cell(row, 2).value = len(self.decryptFile.baseBinList)
                row += 1

                for baseBinName in self.decryptFile.baseBinList:
                    ws.cell(row, 1).value = baseBinName
                    row += 1
                row += 1

            # bin ANIME
            ws.cell(row, 1).value = "binAnime"
            if self.decryptFile.game in ["BS", "CS", "RS"]:
                ws.cell(row, 2).value = len(self.decryptFile.binAnimeList)
            row += 1

            for binAnimeInfo in self.decryptFile.binAnimeList:
                for idx, binAnime in enumerate(binAnimeInfo):
                    ws.cell(row, 1 + idx).value = binAnime
                row += 1
            row += 1
        # smf情報
        elif sheetIndex == 2:
            ws.cell(row, 1).value = "MdlCnt"
            ws.cell(row, 2).value = len(self.decryptFile.smfList)
            row += 1

            for smfIdx, smfInfo in enumerate(self.decryptFile.smfList):
                ws.cell(row, 1).value = smfIdx
                if self.decryptFile.game in ["CS", "RS"]:
                    for idx, smf in enumerate(smfInfo):
                        if idx in [1, 2]:
                            smf = self.toHex(smf)
                        ws.cell(row, 2 + idx).value = smf
                    row += 1
                else:
                    idx = 0
                    for smf in smfInfo[:-1]:
                        ws.cell(row, 2 + idx).value = smf
                        idx += 1
                    ws.cell(row, 2 + idx).value = len(smfInfo[-1])
                    idx += 1
                    animeIdx = idx
                    for smfAnimeInfo in smfInfo[-1]:
                        idx = animeIdx
                        for smfAnime in smfAnimeInfo:
                            ws.cell(row, 2 + idx).value = smfAnime
                            idx += 1
                        row += 1
                    if len(smfInfo[-1]) == 0:
                        row += 1
            row += 1
        # 駅名位置情報
        elif sheetIndex == 3:
            ws.cell(row, 1).value = "STCnt"
            ws.cell(row, 2).value = len(self.decryptFile.stationNameList)
            row += 1

            for stIdx, stationNameInfo in enumerate(self.decryptFile.stationNameList):
                ws.cell(row, 1).value = stIdx
                for idx, stInfo in enumerate(stationNameInfo):
                    ws.cell(row, 2 + idx).value = stInfo
                row += 1
            row += 1
        # 要素２
        elif sheetIndex == 4:
            ws.cell(row, 1).value = "else2"
            ws.cell(row, 2).value = len(self.decryptFile.else2List)
            row += 1

            for else2Info in self.decryptFile.else2List:
                for idx, else2 in enumerate(else2Info):
                    ws.cell(row, 1 + idx).value = else2
                row += 1
            row += 1
        # CPU情報
        elif sheetIndex == 5:
            ws.cell(row, 1).value = "CPU"
            ws.cell(row, 2).value = len(self.decryptFile.cpuList)
            row += 1

            for cpuIdx, cpuInfo in enumerate(self.decryptFile.cpuList):
                ws.cell(row, 1).value = cpuIdx
                for idx, cpu in enumerate(cpuInfo):
                    if self.decryptFile.game == "LS":
                        if idx == 0:
                            ws.cell(row, 2 + idx).value = cpu
                            row += 1
                        elif idx == 1:
                            for list1Idx, list1Info in enumerate(cpu):
                                ws.cell(row, 3 + list1Idx).value = list1Info
                            row -= 1
                        elif idx == 9:
                            for list2Idx, list2Info in enumerate(cpu):
                                ws.cell(row, 3 + list2Idx).value = list2Info
                        else:
                            ws.cell(row, 1 + idx).value = cpu
                            if idx == 8:
                                row += 2
                    else:
                        ws.cell(row, 2 + idx).value = cpu
                row += 1
            row += 1
        # Comic Script、土讃線
        elif sheetIndex == 6:
            ws.cell(row, 1).value = "ComicScript"
            ws.cell(row, 2).value = len(self.decryptFile.comicScriptList)
            row += 1

            for scriptIdx, comicScriptInfo in enumerate(self.decryptFile.comicScriptList):
                ws.cell(row, 1).value = scriptIdx
                if self.decryptFile.game == "LS":
                    for idx, comicScript in enumerate(comicScriptInfo[:-1]):
                        ws.cell(row, 2 + idx).value = comicScript
                    row += 1

                    for idx, script in enumerate(comicScriptInfo[-1]):
                        ws.cell(row, 2 + idx).value = script
                else:
                    for idx, comicScript in enumerate(comicScriptInfo):
                        ws.cell(row, 2 + idx).value = comicScript
                row += 1
            row += 1

            if self.decryptFile.game in ["CS", "RS"]:
                ws.cell(row, 1).value = "DosanInfo"
                ws.cell(row, 2).value = len(self.decryptFile.dosansenList)
                row += 1

                for dosanIdx, dosansenInfo in enumerate(self.decryptFile.dosansenList):
                    ws.cell(row, 1).value = dosanIdx
                    for idx, dosansen in enumerate(dosansenInfo):
                        if idx <= 5:
                            ws.cell(row, 2 + idx).value = dosansen
                            if idx == 5:
                                row += 1
                        else:
                            ws.cell(row, idx - 4).value = dosansen
                    row += 1
                row += 1
        # レール情報
        elif sheetIndex == 7:
            ws.cell(row, 1).value = "RailCnt"
            ws.cell(row, 2).value = len(self.decryptFile.railList)
            row += 2

            titleList = [
                "index",
                "prev_rail",
                "block",
                "dir_x",
                "dir_y",
                "dir_z",
                "mdl_no",
                "mdl_kasen",
                "mdl_kasenchu",
                "per",
                "flg",
                "flg",
                "flg",
                "flg",
                "rail_data",
                "next_rail",
                "next_no",
                "prev_rail",
                "prev_no",
            ]
            mdlList = [x[0] for x in self.decryptFile.smfList]
            readFlag = False
            if self.decryptFile.game == "LS":
                if self.decryptFile.ver == "DEND_MAP_VER0101":
                    readFlag = True
                titleList.pop(2)
                titleList.insert(8, "fix_amb_mdl")
                titleList.insert(8, "rot_z")
                titleList.insert(8, "rot_y")
                titleList.insert(8, "rot_x")
                titleList.insert(2, "pos_z")
                titleList.insert(2, "pos_y")
                titleList.insert(2, "pos_x")
                idx = 0
                for title in titleList:
                    ws.cell(row, 1 + idx).value = title
                    idx += 1
                row += 1

                idx = 0
                prevRail = -1
                offset = 0
                if readFlag:
                    offset = 2

                for railInfo in self.decryptFile.railList:
                    idx = 0
                    for railIdx, rail in enumerate(railInfo[:-1]):
                        if readFlag:
                            if railIdx in [1, 2]:
                                continue
                        if railIdx == 0:
                            ws.cell(row, 1 + idx).value = rail
                            idx += 1
                        elif railIdx >= 1 + offset and railIdx <= 6 + offset:
                            ws.cell(row, 2 + idx).value = rail
                            idx += 1
                        # mdl_no
                        elif railIdx == 7 + offset:
                            rail = self.getSmfModelName(rail, mdlList)
                            ws.cell(row, 2 + idx).value = rail
                            idx += 1
                        # prevRail
                        elif railIdx == 8 + offset:
                            ws.cell(row, 2).value = rail
                            prevRail = rail
                            idx += 1
                        # rot
                        elif railIdx >= 9 + offset and railIdx <= 11 + offset and prevRail == -1:
                            ws.cell(row, 3 + idx).value = rail
                            idx += 1
                        # kasenchu
                        elif (railIdx == 12 + offset and prevRail == -1) or (railIdx == 9 + offset and prevRail != -1):
                            rail = self.getSmfModelName(rail, mdlList)
                            ws.cell(row, 11).value = rail
                            idx += 1
                        # kasen
                        elif (railIdx == 13 + offset and prevRail == -1) or (railIdx == 10 + offset and prevRail != -1):
                            rail = self.getSmfModelName(rail, mdlList)
                            ws.cell(row, 10).value = rail
                            if prevRail != -1:
                                idx += 3
                        # fix_amb_mdl
                        elif (railIdx == 14 + offset and prevRail == -1) or (railIdx == 11 + offset and prevRail != -1):
                            rail = self.getSmfModelName(rail, mdlList)
                            ws.cell(row, 2 + idx).value = rail
                            idx += 1
                        # flg
                        elif (railIdx >= 16 + offset and railIdx <= 19 + offset and prevRail == -1) or (railIdx >= 13 + offset and railIdx <= 16 + offset and prevRail != -1):
                            ws.cell(row, 2 + idx).value = self.toHex(rail)
                            idx += 1
                        else:
                            ws.cell(row, 2 + idx).value = rail
                            idx += 1
                    row += 1
            elif self.decryptFile.game == "BS":
                idx = 0
                for title in titleList:
                    ws.cell(row, 1 + idx).value = title
                    idx += 1
                row += 1

                for railInfo in self.decryptFile.railList:
                    idx = 0
                    for railIdx, rail in enumerate(railInfo[:-3]):
                        # mdl_no, kasen, kasenchu
                        if railIdx >= 6 and railIdx <= 8:
                            rail = self.getSmfModelName(rail, mdlList)
                        # flg
                        if railIdx >= 10 and railIdx <= 13:
                            rail = self.toHex(rail)
                        ws.cell(row, 1 + idx).value = rail
                        idx += 1
                    row += 1
            elif self.decryptFile.game == "CS":
                idx = 0
                for title in titleList:
                    ws.cell(row, 1 + idx).value = title
                    idx += 1
                row += 1

                for railInfo in self.decryptFile.railList:
                    idx = 0
                    for railIdx, rail in enumerate(railInfo[:-2]):
                        # mdl_no, kasen, kasenchu
                        if railIdx >= 6 and railIdx <= 8:
                            rail = self.getSmfModelName(rail, mdlList)
                        # flg
                        if railIdx >= 10 and railIdx <= 13:
                            rail = self.toHex(rail)
                        ws.cell(row, 1 + idx).value = rail
                        idx += 1
                    row += 1
            elif self.decryptFile.game == "RS":
                idx = 0
                for title in titleList:
                    ws.cell(row, 1 + idx).value = title
                    idx += 1
                row += 1

                for railInfo in self.decryptFile.railList:
                    idx = 0
                    for railIdx, rail in enumerate(railInfo):
                        # mdl_no, kasen, kasenchu
                        if railIdx >= 6 and railIdx <= 8:
                            rail = self.getSmfModelName(rail, mdlList)
                        # flg
                        if railIdx >= 10 and railIdx <= 13:
                            rail = self.toHex(rail)
                        ws.cell(row, 1 + idx).value = rail
                        idx += 1
                    row += 1

        # 要素３（Cam）
        elif sheetIndex == 8:
            ws.cell(row, 1).value = "else3"
            ws.cell(row, 2).value = len(self.decryptFile.else3List)
            row += 1

            for else3Idx, else3Info in enumerate(self.decryptFile.else3List):
                ws.cell(row, 1).value = else3Idx
                idx = 0
                for else3 in else3Info[:-1]:
                    ws.cell(row, 2 + idx).value = else3
                    idx += 1
                ws.cell(row, 2 + idx).value = len(else3Info[-1])
                idx += 1
                listIdx = idx
                for listInfo in else3Info[-1]:
                    idx = listIdx
                    for e3list in listInfo:
                        ws.cell(row, 2 + idx).value = e3list
                        idx += 1
                    row += 1
                if len(else3Info[-1]) == 0:
                    row += 1
        # 要素４
        elif sheetIndex == 9:
            ws.cell(row, 1).value = "else4"
            ws.cell(row, 2).value = len(self.decryptFile.else4List)
            row += 1

            for else4Idx, else4Info in enumerate(self.decryptFile.else4List):
                ws.cell(row, 1).value = else4Idx
                for idx, else4 in enumerate(else4Info):
                    ws.cell(row, 2 + idx).value = else4
                    idx += 1
                row += 1
            row += 1
        # AMB情報
        elif sheetIndex == 10:
            ws.cell(row, 1).value = "AmbCnt"
            ws.cell(row, 2).value = len(self.decryptFile.ambList)
            row += 1
            mdlList = [x[0] for x in self.decryptFile.smfList]

            if self.decryptFile.game in ["LS", "BS"]:
                for ambIdx, ambInfo in enumerate(self.decryptFile.ambList):
                    ws.cell(row, 1).value = ambIdx
                    for idx, amb in enumerate(ambInfo):
                        ws.cell(row, 2 + idx).value = amb
                        idx += 1
                    row += 1
                row += 1
            elif self.decryptFile.game in ["CS", "RS"]:
                row += 1
                titleList = [
                    "index",
                    "type",
                    "length",
                    "rail_no",
                    "rail_pos",
                    "base_pos_x",
                    "base_pos_y",
                    "base_pos_z",
                    "base_dir_x",
                    "base_dir_y",
                    "base_dir_z",
                    "priority",
                    "fog|child count",
                    "mdl_no",
                    "pos_x",
                    "pos_y",
                    "pos_z",
                    "dir_x",
                    "dir_y",
                    "dir_z",
                    "dir_x2",
                    "dir_y2",
                    "dir_z2",
                    "per"
                ]

                idx = 0
                for title in titleList:
                    ws.cell(row, 1 + idx).value = title
                    idx += 1
                row += 1
                for ambIdx, ambInfo in enumerate(self.decryptFile.ambList):
                    ws.cell(row, 1).value = ambIdx
                    idx = 0
                    childFlag = False
                    for amb in ambInfo:
                        if not childFlag:
                            # mdl_no
                            if idx == 12:
                                amb = self.getSmfModelName(amb, mdlList)
                            # child count
                            elif idx == 23:
                                row += 1
                                ws.cell(row, 13).value = amb
                                childFlag = True
                                idx = 12
                                continue

                        if childFlag:
                            if idx == 23:
                                row += 1
                                idx = 12

                            if idx == 12:
                                amb = self.getSmfModelName(amb, mdlList)
                        ws.cell(row, 2 + idx).value = amb
                        idx += 1
                    row += 1
                row += 1

    def toHex(self, num):
        return "0x{:02x}".format(num)

    def save(self):
        filename = self.decryptFile.filename + ".xlsx"
        file_path = fd.askopenfilename(initialfile=filename, defaultextension="xlsx", filetypes=[("railData", "*.xlsx")])
        if not file_path:
            return

        wb = openpyxl.load_workbook(file_path, data_only=True)
        # TabList
        tabList = textSetting.textList["railEditor"]["railComboValue"]
        if self.decryptFile.game == "LS":
            tabList = textSetting.textList["railEditor"]["railLsComboValue"]

        for tabName in tabList:
            if tabName not in wb.sheetnames:
                mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E95"].format(tabName))
                return

        newByteArr = bytearray()
        try:
            if self.decryptFile.game == "LS":
                readFlag = False
                # ver
                ws = wb[tabList[0]]
                ver = ws.cell(1, 1).value
                if ver == "DEND_MAP_VER0101":
                    readFlag = True
                bVer = ver.encode("shift-jis")
                newByteArr.extend(bVer)

                # smf情報
                smfNameList = []

                ws = wb[tabList[2]]
                row = self.findLabel("MdlCnt", ws["A"])
                smfCnt = ws.cell(row, 2).value
                newByteArr.append(smfCnt)
                row += 1
                for i in range(smfCnt):
                    smfName = ws.cell(row, 2).value
                    smfNameList.append(smfName)

                    bSmfName = smfName.encode("shift-jis")
                    newByteArr.append(len(bSmfName))
                    newByteArr.extend(bSmfName)

                    newByteArr.append(ws.cell(row, 3).value)
                    newByteArr.append(ws.cell(row, 4).value)
                    cnt = ws.cell(row, 5).value
                    if cnt == 0:
                        newByteArr.append(0xFF)
                        row += 1
                    else:
                        newByteArr.append(cnt)
                        for j in range(cnt):
                            tempH = struct.pack("<h", ws.cell(row, 6).value)
                            newByteArr.extend(tempH)
                            tempH = struct.pack("<h", ws.cell(row, 7).value)
                            newByteArr.extend(tempH)
                            row += 1

                # BGM
                ws = wb[tabList[0]]
                row = self.findLabel("BGM", ws["A"]) + 1
                musicFile = ws.cell(row, 1).value
                bMusicFile = musicFile.encode("shift-jis")
                newByteArr.append(len(bMusicFile))
                newByteArr.extend(bMusicFile)

                musicName = ws.cell(row, 2).value
                bMusicName = musicName.encode("shift-jis")
                newByteArr.append(len(bMusicName))
                newByteArr.extend(bMusicName)

                start = ws.cell(row, 3).value
                newByteArr.extend(struct.pack("<f", start))
                loopStart = ws.cell(row, 4).value
                newByteArr.extend(struct.pack("<f", loopStart))

                # レール名
                row = self.findLabel("railName", ws["A"])
                railStationName = ws.cell(row, 2).value
                bRailStationName = railStationName.encode("shift-jis")
                newByteArr.append(len(bRailStationName))
                newByteArr.extend(bRailStationName)

                # SCENE 3D OBJ(bin ANIME)
                ws = wb[tabList[1]]
                row = self.findLabel("binAnime", ws["A"]) + 1
                for i in range(3):
                    newByteArr.append(ws.cell(row, 1 + i).value)

                row = self.findLabel("else1", ws["A"])
                cnt = ws.cell(row, 2).value
                newByteArr.append(cnt)
                row += 1
                for i in range(cnt):
                    tempF = struct.pack("<f", ws.cell(row, 1 + i).value)
                    newByteArr.extend(tempF)

                # 車両の初期レール位置
                ws = wb[tabList[0]]
                row = self.findLabel("RailPos", ws["A"])
                trainCnt = ws.cell(row, 2).value
                newByteArr.append(trainCnt)
                row += 1
                for i in range(trainCnt):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                else4Dict = {}
                # else4情報
                if readFlag:
                    ws = wb[tabList[9]]
                    row = self.findLabel("else4", ws["A"])
                    else4Cnt = ws.cell(row, 2).value
                    row += 1
                    for i in range(else4Cnt):
                        railNo = ws.cell(row, 2).value
                        if railNo not in else4Dict:
                            else4Dict[railNo] = []
                            for j in range(7):
                                else4Dict[railNo].append(ws.cell(row, 3 + j).value)
                        row += 1

                # AMB情報
                ws = wb[tabList[10]]
                row = self.findLabel("AmbCnt", ws["A"])
                ambCnt = ws.cell(row, 2).value
                row += 1
                ambDict = {}
                for i in range(ambCnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in ambDict:
                        ambDict[railNo] = []
                    ambInfo = []
                    for j in range(4):
                        ambInfo.append(ws.cell(row, 3 + j).value)
                    ambDict[railNo].append(ambInfo)
                    row += 1

                # レール情報
                dupNum = -1
                dupName = None
                ws = wb[tabList[7]]
                row = self.findLabel("RailCnt", ws["A"])
                railCnt = ws.cell(row, 2).value
                hRailCnt = struct.pack("<h", railCnt)
                newByteArr.extend(hRailCnt)
                row = self.findLabel("index", ws["A"]) + 1
                for i in range(railCnt):
                    if readFlag:
                        if i in else4Dict:
                            prevRail2 = else4Dict[i][0]
                            hPrevRail2 = struct.pack("<h", prevRail2)
                            newByteArr.extend(hPrevRail2)
                            for else4 in else4Dict[i][1:]:
                                tempF = struct.pack("<f", else4)
                                newByteArr.extend(tempF)
                        else:
                            tempH = struct.pack("<h", -1)
                            newByteArr.extend(tempH)

                    # pos, dir
                    for j in range(6):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)

                    # mdl_no
                    mdl_no = ws.cell(row, 9).value
                    if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 9).value
                    mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                    newByteArr.append(mdl_no)

                    # prevRail
                    prevRail = ws.cell(row, 2).value
                    hPrevRail = struct.pack("<h", prevRail)
                    newByteArr.extend(hPrevRail)

                    # rot
                    if prevRail == -1:
                        for j in range(3):
                            tempF = struct.pack("<f", ws.cell(row, 12 + j).value)
                            newByteArr.extend(tempF)

                    # 架線柱
                    kasenchu = ws.cell(row, 11).value
                    if self.isModelNameDup(kasenchu, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 11).value
                    kasenchu = self.getSmfModelIndex(i, kasenchu, smfNameList)
                    if kasenchu > 127:
                        bKasenchu = struct.pack("<B", kasenchu)
                    else:
                        bKasenchu = struct.pack("<b", kasenchu)
                    newByteArr.extend(bKasenchu)

                    # 架線
                    kasen = ws.cell(row, 10).value
                    if self.isModelNameDup(kasen, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 10).value
                    kasen = self.getSmfModelIndex(i, kasen, smfNameList)
                    if kasen > 127:
                        bKasen = struct.pack("<B", kasen)
                    else:
                        bKasen = struct.pack("<b", kasen)
                    newByteArr.extend(bKasen)

                    # dummy?
                    for j in range(2):
                        newByteArr.append(0xFF)
                        for k in range(3):
                            tempF = struct.pack("<f", 0)
                            newByteArr.extend(tempF)

                    # fix_amb_mdl
                    fix_amb_mdl = ws.cell(row, 15).value
                    if self.isModelNameDup(fix_amb_mdl, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 15).value
                    fix_amb_mdl = self.getSmfModelIndex(i, fix_amb_mdl, smfNameList)
                    if fix_amb_mdl > 127:
                        bFixAmb = struct.pack("<B", fix_amb_mdl)
                    else:
                        bFixAmb = struct.pack("<b", fix_amb_mdl)
                    newByteArr.extend(bFixAmb)

                    # per
                    per = ws.cell(row, 16).value
                    perF = struct.pack("<f", per)
                    newByteArr.extend(perF)

                    # flg
                    for j in range(4):
                        flg = ws.cell(row, 17 + j).value
                        flg = int(flg, 16)
                        newByteArr.append(flg)

                    # raildata
                    raildata = ws.cell(row, 21).value
                    newByteArr.append(raildata)
                    for j in range(raildata):
                        nextRailNo = ws.cell(row, 22 + 4*j).value
                        hNextRailNo = struct.pack("<h", nextRailNo)
                        newByteArr.extend(hNextRailNo)
                        nextRailPos = ws.cell(row, 23 + 4*j).value
                        hNextRailPos = struct.pack("<h", nextRailPos)
                        newByteArr.extend(hNextRailPos)
                        prevRailNo = ws.cell(row, 24 + 4*j).value
                        hPrevRailNo = struct.pack("<h", prevRailNo)
                        newByteArr.extend(hPrevRailNo)
                        prevRailPos = ws.cell(row, 25 + 4*j).value
                        hPrevRailPos = struct.pack("<h", prevRailPos)
                        newByteArr.extend(hPrevRailPos)

                    # AMB情報
                    if i in ambDict:
                        ambList = ambDict[i]
                        newByteArr.append(len(ambList))
                        for ambInfo in ambList:
                            pos = ambInfo[0]
                            newByteArr.append(pos)

                            railPos = ambInfo[1]
                            railPosH = struct.pack("<h", railPos)
                            newByteArr.extend(railPosH)

                            anime1 = ambInfo[2]
                            newByteArr.append(anime1)

                            anime2 = ambInfo[3]
                            bAnime2 = struct.pack("<b", anime2)
                            newByteArr.extend(bAnime2)
                    else:
                        newByteArr.append(0)
                    row += 1

                if dupNum != -1:
                    warnMsg = textSetting.textList["infoList"]["I115"].format(dupNum, dupName)
                    result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
                    if not result:
                        return

                # 駅名位置情報
                ws = wb[tabList[3]]
                row = self.findLabel("STCnt", ws["A"])
                stCnt = ws.cell(row, 2).value
                newByteArr.append(stCnt)
                row += 1

                for i in range(stCnt):
                    stName = ws.cell(row, 2).value
                    bStName = stName.encode("shift-jis")
                    newByteArr.append(len(bStName))
                    newByteArr.extend(bStName)

                    stFlag = ws.cell(row, 3).value
                    newByteArr.append(stFlag)
                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    for j in range(6):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    row += 1

                # Cam
                ws = wb[tabList[8]]
                row = self.findLabel("else3", ws["A"])
                else3Cnt = ws.cell(row, 2).value
                newByteArr.append(else3Cnt)
                row += 1
                for i in range(else3Cnt):
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 2 + j).value)
                        newByteArr.extend(tempF)
                    cnt = ws.cell(row, 5).value
                    newByteArr.append(cnt)

                    for j in range(cnt):
                        for k in range(4):
                            tempF = struct.pack("<f", ws.cell(row, 6 + k).value)
                            newByteArr.extend(tempF)
                        cameraNo = ws.cell(row, 10).value
                        newByteArr.append(cameraNo)
                        row += 1
                    if cnt == 0:
                        row += 1

                # CPU情報
                ws = wb[tabList[5]]
                row = self.findLabel("CPU", ws["A"])
                cpuCnt = ws.cell(row, 2).value
                newByteArr.append(cpuCnt)
                row += 1
                for i in range(cpuCnt):
                    railNo = ws.cell(row, 2).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    row += 1

                    for j in range(6):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)

                    row -= 1
                    org = ws.cell(row, 3).value
                    newByteArr.append(org)
                    mode = ws.cell(row, 4).value
                    newByteArr.append(mode)

                    for j in range(5):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    row += 2

                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)
                    row += 1

                # Comic Script
                ws = wb[tabList[6]]
                row = self.findLabel("ComicScript", ws["A"])
                comicbinCnt = ws.cell(row, 2).value
                newByteArr.append(comicbinCnt)
                row += 1
                for i in range(comicbinCnt):
                    comicNum = ws.cell(row, 2).value
                    hComicNum = struct.pack("<h", comicNum)
                    newByteArr.extend(hComicNum)

                    comicType = ws.cell(row, 3).value
                    newByteArr.append(comicType)

                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    row += 1

                    for j in range(9):
                        tempF = struct.pack("<f", ws.cell(row, 2 + j).value)
                        newByteArr.extend(tempF)
                    row += 1
            elif self.decryptFile.game == "BS":
                # ver
                ws = wb[tabList[0]]
                ver = ws.cell(1, 1).value
                bVer = ver.encode("shift-jis")
                newByteArr.extend(bVer)

                # レール名
                row = self.findLabel("railName", ws["A"])
                railStationName = ws.cell(row, 2).value
                bRailStationName = railStationName.encode("shift-jis")
                newByteArr.append(len(bRailStationName))
                newByteArr.extend(bRailStationName)

                # BGM
                row = self.findLabel("BGM", ws["A"])
                musicCnt = ws.cell(row, 2).value
                newByteArr.append(musicCnt)
                row += 1
                for i in range(musicCnt):
                    musicFile = ws.cell(row, 1).value
                    bMusicFile = musicFile.encode("shift-jis")
                    newByteArr.append(len(bMusicFile))
                    newByteArr.extend(bMusicFile)

                    musicName = ws.cell(row, 2).value
                    bMusicName = musicName.encode("shift-jis")
                    newByteArr.append(len(bMusicName))
                    newByteArr.extend(bMusicName)

                    start = ws.cell(row, 3).value
                    newByteArr.extend(struct.pack("<f", start))
                    loopStart = ws.cell(row, 4).value
                    newByteArr.extend(struct.pack("<f", loopStart))
                    row += 1

                # 車両の初期レール位置
                row = self.findLabel("RailPos", ws["A"])
                trainCnt = ws.cell(row, 2).value
                newByteArr.append(trainCnt)
                row += 1
                for i in range(trainCnt):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                # ダミー位置？
                row = self.findLabel("RailPos2", ws["A"]) + 1
                railNo = ws.cell(row, 1).value
                hRailNo = struct.pack("<h", railNo)
                newByteArr.extend(hRailNo)
                railPos = ws.cell(row, 2).value
                hRailPos = struct.pack("<h", railPos)
                newByteArr.extend(hRailPos)
                newByteArr.append(ws.cell(row, 3).value)
                f1 = ws.cell(row, 4).value
                tempF = struct.pack("<f", f1)
                newByteArr.extend(tempF)

                # 試運転、二人バトルの初期レール位置
                row = self.findLabel("FreeRunOrVSPos", ws["A"]) + 1
                for i in range(2):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                # 駅表示を始める番号
                row = self.findLabel("stationNo", ws["A"])
                newByteArr.append(ws.cell(row, 2).value)

                row = self.findLabel("RailPos4", ws["A"]) + 1
                railNo = ws.cell(row, 1).value
                hRailNo = struct.pack("<h", railNo)
                newByteArr.extend(hRailNo)
                railPos = ws.cell(row, 2).value
                hRailPos = struct.pack("<h", railPos)
                newByteArr.extend(hRailPos)
                newByteArr.append(ws.cell(row, 3).value)
                f1 = ws.cell(row, 4).value
                tempF = struct.pack("<f", f1)
                newByteArr.extend(tempF)

                row = self.findLabel("stationNo2", ws["A"])
                newByteArr.append(ws.cell(row, 2).value)

                # 要素１
                ws = wb[tabList[1]]
                row = self.findLabel("else1-1", ws["A"])
                tempF = struct.pack("<f", ws.cell(row, 2).value)
                newByteArr.extend(tempF)

                row = self.findLabel("else1-2", ws["A"])
                cnt = ws.cell(row, 2).value
                newByteArr.append(cnt)
                row += 1
                for i in range(cnt):
                    for j in range(2):
                        tempF = struct.pack("<f", ws.cell(row, 1 + j).value)
                        newByteArr.extend(tempF)
                    for j in range(3):
                        newByteArr.append(ws.cell(row, 3 + j).value)
                    row += 1

                row = self.findLabel("light", ws["A"])
                lightCnt = ws.cell(row, 2).value
                newByteArr.append(lightCnt)
                row += 1
                for i in range(lightCnt):
                    lightFile = ws.cell(row, 1).value
                    bLightFile = lightFile.encode("shift-jis")
                    newByteArr.append(len(bLightFile))
                    newByteArr.extend(bLightFile)
                    row += 1

                row = self.findLabel("baseBin", ws["A"])
                baseBinCnt = ws.cell(row, 2).value
                newByteArr.append(baseBinCnt)
                row += 1
                for i in range(baseBinCnt):
                    baseBinFile = ws.cell(row, 1).value
                    bBaseBinFile = baseBinFile.encode("shift-jis")
                    newByteArr.append(len(bBaseBinFile))
                    newByteArr.extend(bBaseBinFile)
                    row += 1

                row = self.findLabel("binAnime", ws["A"])
                binAnimeCnt = ws.cell(row, 2).value
                newByteArr.append(binAnimeCnt)
                row += 1
                for i in range(binAnimeCnt):
                    binIndex = ws.cell(row, 1).value
                    newByteArr.append(binIndex)
                    binAnime1 = ws.cell(row, 2).value
                    hBinAnime1 = struct.pack("<h", binAnime1)
                    newByteArr.extend(hBinAnime1)
                    binAnime2 = ws.cell(row, 3).value
                    hBinAnime2 = struct.pack("<h", binAnime2)
                    newByteArr.extend(hBinAnime2)
                    row += 1

                # smf情報
                smfNameList = []

                ws = wb[tabList[2]]
                row = self.findLabel("MdlCnt", ws["A"])
                smfCnt = ws.cell(row, 2).value
                newByteArr.append(smfCnt)
                row += 1
                for i in range(smfCnt):
                    smfName = ws.cell(row, 2).value
                    smfNameList.append(smfName)

                    bSmfName = smfName.encode("shift-jis")
                    newByteArr.append(len(bSmfName))
                    newByteArr.extend(bSmfName)

                    newByteArr.append(ws.cell(row, 3).value)
                    newByteArr.append(ws.cell(row, 4).value)
                    newByteArr.append(ws.cell(row, 5).value)
                    cnt = ws.cell(row, 6).value
                    newByteArr.append(cnt)
                    if cnt == 0:
                        row += 1
                    else:
                        for j in range(cnt):
                            tempH = struct.pack("<h", ws.cell(row, 7).value)
                            newByteArr.extend(tempH)
                            tempH = struct.pack("<h", ws.cell(row, 8).value)
                            newByteArr.extend(tempH)
                            tempH = struct.pack("<h", ws.cell(row, 9).value)
                            newByteArr.extend(tempH)
                            row += 1

                # 駅名位置情報
                ws = wb[tabList[3]]
                row = self.findLabel("STCnt", ws["A"])
                stCnt = ws.cell(row, 2).value
                newByteArr.append(stCnt)
                row += 1

                for i in range(stCnt):
                    stName = ws.cell(row, 2).value
                    bStName = stName.encode("shift-jis")
                    newByteArr.append(len(bStName))
                    newByteArr.extend(bStName)

                    stFlag = ws.cell(row, 3).value
                    newByteArr.append(stFlag)
                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    row += 1

                # 要素２
                ws = wb[tabList[4]]
                row = self.findLabel("else2", ws["A"])
                else2Cnt = ws.cell(row, 2).value
                newByteArr.append(else2Cnt)
                row += 1
                for i in range(else2Cnt):
                    for j in range(2):
                        tempH = struct.pack("<h", ws.cell(row, 1 + j).value)
                        newByteArr.extend(tempH)
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)
                    newByteArr.append(ws.cell(row, 6).value)
                    row += 1

                # CPU情報
                ws = wb[tabList[5]]
                row = self.findLabel("CPU", ws["A"])
                cpuCnt = ws.cell(row, 2).value
                newByteArr.append(cpuCnt)
                row += 1
                for i in range(cpuCnt):
                    railNo = ws.cell(row, 2).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)

                    org = ws.cell(row, 3).value
                    newByteArr.append(org)
                    mode = ws.cell(row, 4).value
                    newByteArr.append(mode)

                    for j in range(4):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    row += 1

                # Comic Script
                ws = wb[tabList[6]]
                row = self.findLabel("ComicScript", ws["A"])
                comicbinCnt = ws.cell(row, 2).value
                newByteArr.append(comicbinCnt)
                row += 1
                for i in range(comicbinCnt):
                    comicNum = ws.cell(row, 2).value
                    hComicNum = struct.pack("<h", comicNum)
                    newByteArr.extend(hComicNum)

                    comicType = ws.cell(row, 3).value
                    newByteArr.append(comicType)

                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    row += 1

                # 要素４
                else4Dict = {}
                ws = wb[tabList[9]]
                row = self.findLabel("else4", ws["A"])
                else4Cnt = ws.cell(row, 2).value
                row += 1
                for i in range(else4Cnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in else4Dict:
                        else4Dict[railNo] = []
                        for j in range(7):
                            else4Dict[railNo].append(ws.cell(row, 3 + j).value)
                    row += 1

                # AMB情報
                ws = wb[tabList[10]]
                row = self.findLabel("AmbCnt", ws["A"])
                ambCnt = ws.cell(row, 2).value
                row += 1
                ambDict = {}
                for i in range(ambCnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in ambDict:
                        ambDict[railNo] = []
                    ambInfo = []
                    for j in range(11):
                        ambInfo.append(ws.cell(row, 3 + j).value)
                    ambDict[railNo].append(ambInfo)
                    row += 1

                # 要素３
                else3Dict = {}
                ws = wb[tabList[8]]
                row = self.findLabel("else3", ws["A"])
                else3Cnt = ws.cell(row, 2).value
                row += 1
                for i in range(else3Cnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in else3Dict:
                        else3Dict[railNo] = []
                    cnt = ws.cell(row, 3).value

                    for j in range(cnt):
                        else3Info = []
                        for k in range(5):
                            else3Info.append(ws.cell(row, 4 + k).value)
                        else3Dict[railNo].append(else3Info)
                        row += 1

                # レール情報
                dupNum = -1
                dupName = None
                ws = wb[tabList[7]]
                row = self.findLabel("RailCnt", ws["A"])
                railCnt = ws.cell(row, 2).value
                hRailCnt = struct.pack("<h", railCnt)
                newByteArr.extend(hRailCnt)
                row = self.findLabel("index", ws["A"]) + 1
                for i in range(railCnt):
                    isDisableFlg = int(ws.cell(row, 14).value, 16) & 128 > 0
                    # prevRail
                    prevRail = ws.cell(row, 2).value
                    hPrevRail = struct.pack("<h", prevRail)
                    newByteArr.extend(hPrevRail)

                    if prevRail == -1:
                        if i in else4Dict:
                            prevRail2 = else4Dict[i][0]
                            hPrevRail2 = struct.pack("<h", prevRail2)
                            newByteArr.extend(hPrevRail2)
                            for else4 in else4Dict[i][1:]:
                                tempF = struct.pack("<f", else4)
                                newByteArr.extend(tempF)
                        else:
                            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E97"].format(i))
                            return

                    # block
                    newByteArr.append(ws.cell(row, 3).value)

                    # dir
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 4 + j).value)
                        newByteArr.extend(tempF)

                    # mdl_no
                    mdl_no = ws.cell(row, 7).value
                    if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 7).value
                    mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                    newByteArr.append(mdl_no)

                    # 架線
                    kasen = ws.cell(row, 8).value
                    if self.isModelNameDup(kasen, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 8).value
                    kasen = self.getSmfModelIndex(i, kasen, smfNameList)
                    if kasen > 127:
                        bKasen = struct.pack("<B", kasen)
                    else:
                        bKasen = struct.pack("<b", kasen)
                    newByteArr.extend(bKasen)

                    # 架線柱
                    kasenchu = ws.cell(row, 9).value
                    if self.isModelNameDup(kasenchu, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 9).value
                    kasenchu = self.getSmfModelIndex(i, kasenchu, smfNameList)
                    if kasenchu > 127:
                        bKasenchu = struct.pack("<B", kasenchu)
                    else:
                        bKasenchu = struct.pack("<b", kasenchu)
                    newByteArr.extend(bKasenchu)

                    # per
                    per = ws.cell(row, 10).value
                    perF = struct.pack("<f", per)
                    newByteArr.extend(perF)

                    # flg
                    for j in range(4):
                        flg = ws.cell(row, 11 + j).value
                        flg = int(flg, 16)
                        newByteArr.append(flg)

                    # raildata
                    raildata = ws.cell(row, 15).value
                    newByteArr.append(raildata)
                    for j in range(raildata):
                        nextRailNo = ws.cell(row, 16 + 4*j).value
                        hNextRailNo = struct.pack("<h", nextRailNo)
                        newByteArr.extend(hNextRailNo)
                        nextRailPos = ws.cell(row, 17 + 4*j).value
                        hNextRailPos = struct.pack("<h", nextRailPos)
                        newByteArr.extend(hNextRailPos)
                        prevRailNo = ws.cell(row, 18 + 4*j).value
                        hPrevRailNo = struct.pack("<h", prevRailNo)
                        newByteArr.extend(hPrevRailNo)
                        prevRailPos = ws.cell(row, 19 + 4*j).value
                        hPrevRailPos = struct.pack("<h", prevRailPos)
                        newByteArr.extend(hPrevRailPos)

                    # AMB情報
                    if i in ambDict:
                        ambList = ambDict[i]
                        newByteArr.append(len(ambList))
                        for ambInfo in ambList:
                            for j in range(4):
                                newByteArr.append(ambInfo[j])

                            for j in range(7):
                                tempF = struct.pack("<f", ambInfo[4 + j])
                                newByteArr.extend(tempF)
                    else:
                        newByteArr.append(0)

                    # else3情報
                    if i in else3Dict:
                        else3List = else3Dict[i]
                        newByteArr.append(len(else3List))
                        for else3Info in else3List:
                            pos = else3Info[0]
                            newByteArr.append(pos)
                            railNo = else3Info[1]
                            hRailNo = struct.pack("<h", railNo)
                            newByteArr.extend(hRailNo)
                            binIndex = else3Info[2]
                            newByteArr.append(binIndex)

                            anime1 = else3Info[3]
                            hAnime1 = struct.pack("<h", anime1)
                            newByteArr.extend(hAnime1)
                            anime2 = else3Info[4]
                            hAnime2 = struct.pack("<h", anime2)
                            newByteArr.extend(hAnime2)
                    else:
                        newByteArr.append(0)
                    row += 1

                if dupNum != -1:
                    warnMsg = textSetting.textList["infoList"]["I115"].format(dupNum, dupName)
                    result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
                    if not result:
                        return
            elif self.decryptFile.game == "CS":
                # ver
                ws = wb[tabList[0]]
                ver = ws.cell(1, 1).value
                bVer = ver.encode("shift-jis")
                newByteArr.extend(bVer)

                # BGM
                row = self.findLabel("BGM", ws["A"])
                musicCnt = ws.cell(row, 2).value
                newByteArr.append(musicCnt)
                row += 1
                for i in range(musicCnt):
                    newByteArr.append(i)

                # 車両の初期レール位置
                row = self.findLabel("RailPos", ws["A"])
                trainCnt = ws.cell(row, 2).value
                newByteArr.append(trainCnt)
                row += 1
                for i in range(trainCnt):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                # ダミー位置？
                row = self.findLabel("RailPos2", ws["A"]) + 1
                railNo = ws.cell(row, 1).value
                hRailNo = struct.pack("<h", railNo)
                newByteArr.extend(hRailNo)
                railPos = ws.cell(row, 2).value
                hRailPos = struct.pack("<h", railPos)
                newByteArr.extend(hRailPos)
                newByteArr.append(ws.cell(row, 3).value)
                f1 = ws.cell(row, 4).value
                tempF = struct.pack("<f", f1)
                newByteArr.extend(tempF)

                # 試運転、二人バトルの初期レール位置
                row = self.findLabel("FreeRunOrVSPos", ws["A"]) + 1
                for i in range(2):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                # 駅表示を始める番号
                row = self.findLabel("stationNo", ws["A"])
                newByteArr.append(ws.cell(row, 2).value)

                # 要素１
                ws = wb[tabList[1]]
                row = self.findLabel("else1-1", ws["A"])
                tempF = struct.pack("<f", ws.cell(row, 2).value)
                newByteArr.extend(tempF)

                row = self.findLabel("else1-2", ws["A"])
                cnt = ws.cell(row, 2).value
                newByteArr.append(cnt)
                row += 1
                for i in range(cnt):
                    for j in range(2):
                        tempF = struct.pack("<f", ws.cell(row, 1 + j).value)
                        newByteArr.extend(tempF)
                    for j in range(3):
                        newByteArr.append(ws.cell(row, 3 + j).value)
                    row += 1

                row = self.findLabel("light", ws["A"])
                lightCnt = ws.cell(row, 2).value
                newByteArr.append(lightCnt)
                row += 1
                for i in range(lightCnt):
                    lightFile = ws.cell(row, 1).value
                    bLightFile = lightFile.encode("shift-jis")
                    newByteArr.append(len(bLightFile))
                    newByteArr.extend(bLightFile)
                    row += 1

                row = self.findLabel("StageRes", ws["A"])
                stageResCnt = ws.cell(row, 2).value
                hStageResCnt = struct.pack("<h", stageResCnt)
                newByteArr.extend(hStageResCnt)
                row += 1
                for i in range(stageResCnt):
                    stageFile = ws.cell(row, 1).value
                    bStageFile = stageFile.encode("shift-jis")
                    newByteArr.append(len(bStageFile))
                    newByteArr.extend(bStageFile)
                    row += 1

                row = self.findLabel("SetTexInfo", ws["A"])
                stageTexCnt = ws.cell(row, 2).value
                hStageTexCnt = struct.pack("<h", stageTexCnt)
                newByteArr.extend(hStageTexCnt)
                row += 1
                for i in range(stageTexCnt):
                    newByteArr.append(ws.cell(row, 1).value)
                    for j in range(4):
                        tempH = struct.pack("<h", ws.cell(row, 2 + j).value)
                        newByteArr.extend(tempH)
                    row += 1

                row = self.findLabel("baseBin", ws["A"])
                baseBinCnt = ws.cell(row, 2).value
                newByteArr.append(baseBinCnt)
                row += 1
                for i in range(baseBinCnt):
                    baseBinFile = ws.cell(row, 1).value
                    bBaseBinFile = baseBinFile.encode("shift-jis")
                    newByteArr.append(len(bBaseBinFile))
                    newByteArr.extend(bBaseBinFile)
                    row += 1

                row = self.findLabel("binAnime", ws["A"])
                binAnimeCnt = ws.cell(row, 2).value
                newByteArr.append(binAnimeCnt)
                row += 1
                for i in range(binAnimeCnt):
                    binIndex = ws.cell(row, 1).value
                    newByteArr.append(binIndex)
                    binAnime1 = ws.cell(row, 2).value
                    hBinAnime1 = struct.pack("<h", binAnime1)
                    newByteArr.extend(hBinAnime1)
                    binAnime2 = ws.cell(row, 3).value
                    hBinAnime2 = struct.pack("<h", binAnime2)
                    newByteArr.extend(hBinAnime2)
                    row += 1

                # smf情報
                smfNameList = []

                ws = wb[tabList[2]]
                row = self.findLabel("MdlCnt", ws["A"])
                smfCnt = ws.cell(row, 2).value
                newByteArr.append(smfCnt)
                row += 1
                for i in range(smfCnt):
                    smfName = ws.cell(row, 2).value
                    smfNameList.append(smfName)

                    bSmfName = smfName.encode("shift-jis")
                    newByteArr.append(len(bSmfName))
                    newByteArr.extend(bSmfName)

                    for j in range(2):
                        newByteArr.append(int(ws.cell(row, 3 + j).value, 16))
                    newByteArr.append(ws.cell(row, 5).value)
                    newByteArr.append(ws.cell(row, 6).value)
                    newByteArr.append(ws.cell(row, 7).value)

                    kasenchu = ws.cell(row, 8).value
                    newByteArr.append(kasenchu)
                    kasen = ws.cell(row, 9).value
                    hKasen = struct.pack("<h", kasen)
                    newByteArr.extend(hKasen)
                    row += 1

                # 駅名位置情報
                ws = wb[tabList[3]]
                row = self.findLabel("STCnt", ws["A"])
                stCnt = ws.cell(row, 2).value
                newByteArr.append(stCnt)
                row += 1

                for i in range(stCnt):
                    stName = ws.cell(row, 2).value
                    if stName is not None:
                        bStName = stName.encode("shift-jis")
                        newByteArr.append(len(bStName))
                        newByteArr.extend(bStName)
                    else:
                        newByteArr.append(0)

                    stFlag = ws.cell(row, 3).value
                    newByteArr.append(stFlag)
                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)

                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    for j in range(3):
                        tempI = struct.pack("<i", ws.cell(row, 8 + j).value)
                        newByteArr.extend(tempI)
                    tempH = struct.pack("<h", ws.cell(row, 11).value)
                    newByteArr.extend(tempH)
                    row += 1

                # 要素２
                ws = wb[tabList[4]]
                row = self.findLabel("else2", ws["A"])
                else2Cnt = ws.cell(row, 2).value
                newByteArr.append(else2Cnt)
                row += 1
                for i in range(else2Cnt):
                    for j in range(2):
                        tempH = struct.pack("<h", ws.cell(row, 1 + j).value)
                        newByteArr.extend(tempH)
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)
                    newByteArr.append(ws.cell(row, 6).value)
                    row += 1

                # CPU情報
                ws = wb[tabList[5]]
                row = self.findLabel("CPU", ws["A"])
                cpuCnt = ws.cell(row, 2).value
                newByteArr.append(cpuCnt)
                row += 1
                for i in range(cpuCnt):
                    railNo = ws.cell(row, 2).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)

                    org = ws.cell(row, 3).value
                    newByteArr.append(org)
                    mode = ws.cell(row, 4).value
                    newByteArr.append(mode)

                    for j in range(4):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    row += 1

                # Comic Script
                ws = wb[tabList[6]]
                row = self.findLabel("ComicScript", ws["A"])
                comicbinCnt = ws.cell(row, 2).value
                newByteArr.append(comicbinCnt)
                row += 1
                for i in range(comicbinCnt):
                    comicNum = ws.cell(row, 2).value
                    hComicNum = struct.pack("<h", comicNum)
                    newByteArr.extend(hComicNum)

                    comicType = ws.cell(row, 3).value
                    newByteArr.append(comicType)

                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    row += 1

                # 土讃線情報データ
                row = self.findLabel("DosanInfo", ws["A"])
                dosansenCnt = ws.cell(row, 2).value
                newByteArr.append(dosansenCnt)
                row += 1
                for i in range(dosansenCnt):
                    for j in range(6):
                        tempH = struct.pack("<h", ws.cell(row, 2 + j).value)
                        newByteArr.extend(tempH)
                    row += 1

                    tempH = struct.pack("<h", ws.cell(row, 2).value)
                    newByteArr.extend(tempH)
                    for j in range(4):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)
                    tempH = struct.pack("<h", ws.cell(row, 7).value)
                    newByteArr.extend(tempH)
                    tempF = struct.pack("<f", ws.cell(row, 8).value)
                    newByteArr.extend(tempF)
                    row += 1

                # 要素３
                else3Dict = {}
                ws = wb[tabList[8]]
                row = self.findLabel("else3", ws["A"])
                else3Cnt = ws.cell(row, 2).value
                row += 1
                for i in range(else3Cnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in else3Dict:
                        else3Dict[railNo] = []
                    cnt = ws.cell(row, 3).value

                    for j in range(cnt):
                        else3Info = []
                        for k in range(5):
                            else3Info.append(ws.cell(row, 4 + k).value)
                        else3Dict[railNo].append(else3Info)
                        row += 1

                # 要素４
                else4Dict = {}
                ws = wb[tabList[9]]
                row = self.findLabel("else4", ws["A"])
                else4Cnt = ws.cell(row, 2).value
                row += 1
                for i in range(else4Cnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in else4Dict:
                        else4Dict[railNo] = []
                        for j in range(7):
                            else4Dict[railNo].append(ws.cell(row, 3 + j).value)
                    row += 1

                # レール情報
                dupNum = -1
                dupName = None
                ws = wb[tabList[7]]
                row = self.findLabel("RailCnt", ws["A"])
                railCnt = ws.cell(row, 2).value
                hRailCnt = struct.pack("<h", railCnt)
                newByteArr.extend(hRailCnt)
                row = self.findLabel("index", ws["A"]) + 1
                for i in range(railCnt):
                    isDisableFlg = int(ws.cell(row, 14).value, 16) & 128 > 0
                    readFlag = False
                    # prevRail
                    prevRail = ws.cell(row, 2).value
                    hPrevRail = struct.pack("<h", prevRail)
                    newByteArr.extend(hPrevRail)

                    if prevRail == -1:
                        readFlag = True

                    # block
                    newByteArr.append(ws.cell(row, 3).value)

                    # dir
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 4 + j).value)
                        newByteArr.extend(tempF)

                    # mdl_no
                    mdl_no = ws.cell(row, 7).value
                    if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 7).value
                    mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                    newByteArr.append(mdl_no)

                    # 架線
                    kasen = ws.cell(row, 8).value
                    if self.isModelNameDup(kasen, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 8).value
                    kasen = self.getSmfModelIndex(i, kasen, smfNameList)
                    if kasen > 127:
                        bKasen = struct.pack("<B", kasen)
                    else:
                        bKasen = struct.pack("<b", kasen)
                    newByteArr.extend(bKasen)

                    # 架線柱
                    kasenchu = ws.cell(row, 9).value
                    if self.isModelNameDup(kasenchu, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 9).value
                    kasenchu = self.getSmfModelIndex(i, kasenchu, smfNameList)
                    if kasenchu > 127:
                        bKasenchu = struct.pack("<B", kasenchu)
                    else:
                        bKasenchu = struct.pack("<b", kasenchu)
                    newByteArr.extend(bKasenchu)

                    # per
                    per = ws.cell(row, 10).value
                    perF = struct.pack("<f", per)
                    newByteArr.extend(perF)

                    # flg
                    for j in range(4):
                        flg = ws.cell(row, 11 + j).value
                        flg = int(flg, 16)
                        newByteArr.append(flg)

                    # raildata
                    raildata = ws.cell(row, 15).value
                    newByteArr.append(raildata)
                    for j in range(raildata):
                        nextRailNo = ws.cell(row, 16 + 4*j).value
                        hNextRailNo = struct.pack("<h", nextRailNo)
                        newByteArr.extend(hNextRailNo)
                        nextRailPos = ws.cell(row, 17 + 4*j).value
                        hNextRailPos = struct.pack("<h", nextRailPos)
                        newByteArr.extend(hNextRailPos)
                        prevRailNo = ws.cell(row, 18 + 4*j).value
                        hPrevRailNo = struct.pack("<h", prevRailNo)
                        newByteArr.extend(hPrevRailNo)
                        prevRailPos = ws.cell(row, 19 + 4*j).value
                        hPrevRailPos = struct.pack("<h", prevRailPos)
                        newByteArr.extend(hPrevRailPos)

                    # else3情報
                    if i in else3Dict:
                        else3List = else3Dict[i]
                        newByteArr.append(len(else3List))
                        for else3Info in else3List:
                            pos = else3Info[0]
                            newByteArr.append(pos)
                            railNo = else3Info[1]
                            hRailNo = struct.pack("<h", railNo)
                            newByteArr.extend(hRailNo)
                            binIndex = else3Info[2]
                            newByteArr.append(binIndex)

                            anime1 = else3Info[3]
                            hAnime1 = struct.pack("<h", anime1)
                            newByteArr.extend(hAnime1)
                            anime2 = else3Info[4]
                            hAnime2 = struct.pack("<h", anime2)
                            newByteArr.extend(hAnime2)
                    else:
                        newByteArr.append(0)

                    if readFlag:
                        if i in else4Dict:
                            prevRail2 = else4Dict[i][0]
                            hPrevRail2 = struct.pack("<h", prevRail2)
                            newByteArr.extend(hPrevRail2)
                            for else4 in else4Dict[i][1:]:
                                tempF = struct.pack("<f", else4)
                                newByteArr.extend(tempF)
                        else:
                            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E97"].format(i))
                            return
                    row += 1

                if dupNum != -1:
                    warnMsg = textSetting.textList["infoList"]["I115"].format(dupNum, dupName)
                    result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
                    if not result:
                        return

                # AMB情報
                dupNum = -1
                dupName = None
                ws = wb[tabList[10]]
                row = self.findLabel("AmbCnt", ws["A"])
                ambCnt = ws.cell(row, 2).value
                hAmbCnt = struct.pack("<h", ambCnt)
                newByteArr.extend(hAmbCnt)
                row = self.findLabel("index", ws["A"]) + 1
                for i in range(ambCnt):
                    newByteArr.append(ws.cell(row, 2).value)
                    fLength = struct.pack("<f", ws.cell(row, 3).value)
                    newByteArr.extend(fLength)
                    # RailNo, RailPos
                    for j in range(2):
                        tempH = struct.pack("<h", ws.cell(row, 4 + j).value)
                        newByteArr.extend(tempH)

                    # base pos_xyz base rot_xyz
                    for j in range(6):
                        tempF = struct.pack("<f", ws.cell(row, 6 + j).value)
                        newByteArr.extend(tempF)

                    for j in range(2):
                        newByteArr.append(ws.cell(row, 12 + j).value)

                    # mdl_no
                    mdl_no = ws.cell(row, 14).value
                    if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 14).value
                    mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                    hMdlNo = struct.pack("<h", mdl_no)
                    newByteArr.extend(hMdlNo)

                    # pos xyz, dir xyz dir2 xyz
                    for j in range(9):
                        tempF = struct.pack("<f", ws.cell(row, 15 + j).value)
                        newByteArr.extend(tempF)

                    perF = struct.pack("<f", ws.cell(row, 24).value)
                    newByteArr.extend(perF)
                    row += 1

                    childCount = ws.cell(row, 13).value
                    newByteArr.append(childCount)
                    for j in range(childCount):
                        # mdl_no
                        mdl_no = ws.cell(row, 14).value
                        if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                            dupNum = i
                            dupName = ws.cell(row, 14).value
                        mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                        hMdlNo = struct.pack("<h", mdl_no)
                        newByteArr.extend(hMdlNo)

                        # pos xyz, dir xyz dir2 xyz
                        for k in range(9):
                            tempF = struct.pack("<f", ws.cell(row, 15 + k).value)
                            newByteArr.extend(tempF)

                        perF = struct.pack("<f", ws.cell(row, 24).value)
                        newByteArr.extend(perF)
                        row += 1
                    if childCount == 0:
                        row += 1

                if dupNum != -1:
                    warnMsg = textSetting.textList["infoList"]["I116"].format(dupNum, dupName)
                    result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
                    if not result:
                        return

            elif self.decryptFile.game == "RS":
                readFlag = False
                # ver
                ws = wb[tabList[0]]
                ver = ws.cell(1, 1).value
                if ver == "DEND_MAP_VER0400":
                    readFlag = True
                bVer = ver.encode("shift-jis")
                newByteArr.extend(bVer)

                # BGM
                row = self.findLabel("BGM", ws["A"])
                musicCnt = ws.cell(row, 2).value
                newByteArr.append(musicCnt)
                row += 1
                for i in range(musicCnt):
                    newByteArr.append(i)

                # 車両の初期レール位置
                row = self.findLabel("RailPos", ws["A"])
                trainCnt = ws.cell(row, 2).value
                newByteArr.append(trainCnt)
                row += 1
                for i in range(trainCnt):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                # ダミー位置？
                row = self.findLabel("RailPos2", ws["A"]) + 1
                railNo = ws.cell(row, 1).value
                hRailNo = struct.pack("<h", railNo)
                newByteArr.extend(hRailNo)
                railPos = ws.cell(row, 2).value
                hRailPos = struct.pack("<h", railPos)
                newByteArr.extend(hRailPos)
                newByteArr.append(ws.cell(row, 3).value)
                f1 = ws.cell(row, 4).value
                tempF = struct.pack("<f", f1)
                newByteArr.extend(tempF)

                # 試運転、二人バトルの初期レール位置
                row = self.findLabel("FreeRunOrVSPos", ws["A"]) + 1
                for i in range(2):
                    railNo = ws.cell(row, 1).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    railPos = ws.cell(row, 2).value
                    hRailPos = struct.pack("<h", railPos)
                    newByteArr.extend(hRailPos)
                    newByteArr.append(ws.cell(row, 3).value)
                    f1 = ws.cell(row, 4).value
                    tempF = struct.pack("<f", f1)
                    newByteArr.extend(tempF)
                    row += 1

                # 駅表示を始める番号
                row = self.findLabel("stationNo", ws["A"])
                newByteArr.append(ws.cell(row, 2).value)

                # 要素１
                ws = wb[tabList[1]]
                row = self.findLabel("else1-1", ws["A"])
                tempF = struct.pack("<f", ws.cell(row, 2).value)
                newByteArr.extend(tempF)

                row = self.findLabel("else1-2", ws["A"])
                cnt = ws.cell(row, 2).value
                newByteArr.append(cnt)
                row += 1
                for i in range(cnt):
                    for j in range(2):
                        tempF = struct.pack("<f", ws.cell(row, 1 + j).value)
                        newByteArr.extend(tempF)
                    for j in range(3):
                        newByteArr.append(ws.cell(row, 3 + j).value)
                    row += 1

                row = self.findLabel("light", ws["A"])
                lightCnt = ws.cell(row, 2).value
                newByteArr.append(lightCnt)
                row += 1
                for i in range(lightCnt):
                    lightFile = ws.cell(row, 1).value
                    bLightFile = lightFile.encode("shift-jis")
                    newByteArr.append(len(bLightFile))
                    newByteArr.extend(bLightFile)
                    row += 1

                row = self.findLabel("StageRes", ws["A"])
                stageResCnt = ws.cell(row, 2).value
                hStageResCnt = struct.pack("<h", stageResCnt)
                newByteArr.extend(hStageResCnt)
                row += 1
                for i in range(stageResCnt):
                    stageFile = ws.cell(row, 1).value
                    bStageFile = stageFile.encode("shift-jis")
                    newByteArr.append(len(bStageFile))
                    newByteArr.extend(bStageFile)
                    row += 1

                row = self.findLabel("SetTexInfo", ws["A"])
                stageTexCnt = ws.cell(row, 2).value
                hStageTexCnt = struct.pack("<h", stageTexCnt)
                newByteArr.extend(hStageTexCnt)
                row += 1
                for i in range(stageTexCnt):
                    newByteArr.append(ws.cell(row, 1).value)
                    for j in range(4):
                        tempH = struct.pack("<h", ws.cell(row, 2 + j).value)
                        newByteArr.extend(tempH)
                    row += 1

                row = self.findLabel("baseBin", ws["A"])
                baseBinCnt = ws.cell(row, 2).value
                newByteArr.append(baseBinCnt)
                row += 1
                for i in range(baseBinCnt):
                    baseBinFile = ws.cell(row, 1).value
                    bBaseBinFile = baseBinFile.encode("shift-jis")
                    newByteArr.append(len(bBaseBinFile))
                    newByteArr.extend(bBaseBinFile)
                    row += 1

                row = self.findLabel("binAnime", ws["A"])
                binAnimeCnt = ws.cell(row, 2).value
                newByteArr.append(binAnimeCnt)
                row += 1
                for i in range(binAnimeCnt):
                    binIndex = ws.cell(row, 1).value
                    newByteArr.append(binIndex)
                    binAnime1 = ws.cell(row, 2).value
                    hBinAnime1 = struct.pack("<h", binAnime1)
                    newByteArr.extend(hBinAnime1)
                    binAnime2 = ws.cell(row, 3).value
                    hBinAnime2 = struct.pack("<h", binAnime2)
                    newByteArr.extend(hBinAnime2)
                    row += 1

                # smf情報
                smfNameList = []

                ws = wb[tabList[2]]
                row = self.findLabel("MdlCnt", ws["A"])
                smfCnt = ws.cell(row, 2).value
                newByteArr.append(smfCnt)
                row += 1
                for i in range(smfCnt):
                    smfName = ws.cell(row, 2).value
                    smfNameList.append(smfName)

                    bSmfName = smfName.encode("shift-jis")
                    newByteArr.append(len(bSmfName))
                    newByteArr.extend(bSmfName)

                    for j in range(2):
                        newByteArr.append(int(ws.cell(row, 3 + j).value, 16))
                    newByteArr.append(ws.cell(row, 5).value)
                    newByteArr.append(ws.cell(row, 6).value)
                    newByteArr.append(ws.cell(row, 7).value)

                    kasenchu = ws.cell(row, 8).value
                    newByteArr.append(kasenchu)
                    kasen = ws.cell(row, 9).value
                    hKasen = struct.pack("<h", kasen)
                    newByteArr.extend(hKasen)
                    row += 1

                # 駅名位置情報
                ws = wb[tabList[3]]
                row = self.findLabel("STCnt", ws["A"])
                stCnt = ws.cell(row, 2).value
                newByteArr.append(stCnt)
                row += 1

                for i in range(stCnt):
                    stName = ws.cell(row, 2).value
                    if stName is not None:
                        bStName = stName.encode("shift-jis")
                        newByteArr.append(len(bStName))
                        newByteArr.extend(bStName)
                    else:
                        newByteArr.append(0)

                    stFlag = ws.cell(row, 3).value
                    newByteArr.append(stFlag)
                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)

                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    for j in range(3):
                        tempI = struct.pack("<i", ws.cell(row, 8 + j).value)
                        newByteArr.extend(tempI)
                    tempH = struct.pack("<h", ws.cell(row, 11).value)
                    newByteArr.extend(tempH)
                    row += 1

                # 要素２
                ws = wb[tabList[4]]
                row = self.findLabel("else2", ws["A"])
                else2Cnt = ws.cell(row, 2).value
                newByteArr.append(else2Cnt)
                row += 1
                for i in range(else2Cnt):
                    for j in range(2):
                        tempH = struct.pack("<h", ws.cell(row, 1 + j).value)
                        newByteArr.extend(tempH)
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)
                    newByteArr.append(ws.cell(row, 6).value)
                    row += 1

                # CPU情報
                ws = wb[tabList[5]]
                row = self.findLabel("CPU", ws["A"])
                cpuCnt = ws.cell(row, 2).value
                newByteArr.append(cpuCnt)
                row += 1
                for i in range(cpuCnt):
                    railNo = ws.cell(row, 2).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)

                    org = ws.cell(row, 3).value
                    newByteArr.append(org)
                    mode = ws.cell(row, 4).value
                    newByteArr.append(mode)

                    for j in range(4):
                        tempF = struct.pack("<f", ws.cell(row, 5 + j).value)
                        newByteArr.extend(tempF)
                    row += 1

                # Comic Script
                ws = wb[tabList[6]]
                row = self.findLabel("ComicScript", ws["A"])
                comicbinCnt = ws.cell(row, 2).value
                newByteArr.append(comicbinCnt)
                row += 1
                for i in range(comicbinCnt):
                    comicNum = ws.cell(row, 2).value
                    hComicNum = struct.pack("<h", comicNum)
                    newByteArr.extend(hComicNum)

                    comicType = ws.cell(row, 3).value
                    newByteArr.append(comicType)

                    railNo = ws.cell(row, 4).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    row += 1

                # 土讃線情報データ
                row = self.findLabel("DosanInfo", ws["A"])
                dosansenCnt = ws.cell(row, 2).value
                newByteArr.append(dosansenCnt)
                row += 1
                for i in range(dosansenCnt):
                    for j in range(6):
                        tempH = struct.pack("<h", ws.cell(row, 2 + j).value)
                        newByteArr.extend(tempH)
                    row += 1

                    tempH = struct.pack("<h", ws.cell(row, 2).value)
                    newByteArr.extend(tempH)
                    for j in range(4):
                        tempF = struct.pack("<f", ws.cell(row, 3 + j).value)
                        newByteArr.extend(tempF)
                    tempH = struct.pack("<h", ws.cell(row, 7).value)
                    newByteArr.extend(tempH)
                    tempF = struct.pack("<f", ws.cell(row, 8).value)
                    newByteArr.extend(tempF)
                    row += 1

                # 要素４情報
                else4Dict = {}
                ws = wb[tabList[9]]
                row = self.findLabel("else4", ws["A"])
                else4Cnt = ws.cell(row, 2).value
                row += 1
                for i in range(else4Cnt):
                    railNo = ws.cell(row, 2).value
                    if railNo not in else4Dict:
                        else4Dict[railNo] = []
                        for j in range(7):
                            else4Dict[railNo].append(ws.cell(row, 3 + j).value)
                    row += 1

                # レール情報
                dupNum = -1
                dupName = None
                ws = wb[tabList[7]]
                row = self.findLabel("RailCnt", ws["A"])
                railCnt = ws.cell(row, 2).value
                hRailCnt = struct.pack("<h", railCnt)
                newByteArr.extend(hRailCnt)
                row = self.findLabel("index", ws["A"]) + 1
                for i in range(railCnt):
                    isDisableFlg = int(ws.cell(row, 14).value, 16) & 128 > 0
                    # prevRail
                    prevRail = ws.cell(row, 2).value
                    hPrevRail = struct.pack("<h", prevRail)
                    newByteArr.extend(hPrevRail)

                    if prevRail == -1 and i != 0 and not isDisableFlg:
                        if i not in else4Dict:
                            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E97"].format(i))
                            return

                    # block
                    newByteArr.append(ws.cell(row, 3).value)

                    # dir
                    for j in range(3):
                        tempF = struct.pack("<f", ws.cell(row, 4 + j).value)
                        newByteArr.extend(tempF)

                    # mdl_no
                    mdl_no = ws.cell(row, 7).value
                    if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 7).value
                    mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                    newByteArr.append(mdl_no)

                    # 架線
                    kasen = ws.cell(row, 8).value
                    if self.isModelNameDup(kasen, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 8).value
                    kasen = self.getSmfModelIndex(i, kasen, smfNameList)
                    if kasen > 127:
                        bKasen = struct.pack("<B", kasen)
                    else:
                        bKasen = struct.pack("<b", kasen)
                    newByteArr.extend(bKasen)

                    # 架線柱
                    kasenchu = ws.cell(row, 9).value
                    if self.isModelNameDup(kasenchu, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 9).value
                    kasenchu = self.getSmfModelIndex(i, kasenchu, smfNameList)
                    if kasenchu > 127:
                        bKasenchu = struct.pack("<B", kasenchu)
                    else:
                        bKasenchu = struct.pack("<b", kasenchu)
                    newByteArr.extend(bKasenchu)

                    # per
                    per = ws.cell(row, 10).value
                    perF = struct.pack("<f", per)
                    newByteArr.extend(perF)

                    # flg
                    for j in range(4):
                        flg = ws.cell(row, 11 + j).value
                        flg = int(flg, 16)
                        newByteArr.append(flg)

                    # raildata
                    raildata = ws.cell(row, 15).value
                    newByteArr.append(raildata)
                    for j in range(raildata):
                        nextRailNo = ws.cell(row, 16 + 4*j).value
                        hNextRailNo = struct.pack("<h", nextRailNo)
                        newByteArr.extend(hNextRailNo)
                        nextRailPos = ws.cell(row, 17 + 4*j).value
                        hNextRailPos = struct.pack("<h", nextRailPos)
                        newByteArr.extend(hNextRailPos)
                        prevRailNo = ws.cell(row, 18 + 4*j).value
                        hPrevRailNo = struct.pack("<h", prevRailNo)
                        newByteArr.extend(hPrevRailNo)
                        prevRailPos = ws.cell(row, 19 + 4*j).value
                        hPrevRailPos = struct.pack("<h", prevRailPos)
                        newByteArr.extend(hPrevRailPos)

                    if readFlag:
                        for j in range(raildata):
                            nextRailNo = ws.cell(row, 16 + 4*raildata + 4*j).value
                            hNextRailNo = struct.pack("<h", nextRailNo)
                            newByteArr.extend(hNextRailNo)
                            nextRailPos = ws.cell(row, 17 + 4*raildata + 4*j).value
                            hNextRailPos = struct.pack("<h", nextRailPos)
                            newByteArr.extend(hNextRailPos)
                            prevRailNo = ws.cell(row, 18 + 4*raildata + 4*j).value
                            hPrevRailNo = struct.pack("<h", prevRailNo)
                            newByteArr.extend(hPrevRailNo)
                            prevRailPos = ws.cell(row, 19 + 4*raildata + 4*j).value
                            hPrevRailPos = struct.pack("<h", prevRailPos)
                            newByteArr.extend(hPrevRailPos)
                    row += 1

                if dupNum != -1:
                    warnMsg = textSetting.textList["infoList"]["I115"].format(dupNum, dupName)
                    result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
                    if not result:
                        return

                # 要素３
                ws = wb[tabList[8]]
                row = self.findLabel("else3", ws["A"])
                else3Cnt = ws.cell(row, 2).value
                hElse3Cnt = struct.pack("<h", else3Cnt)
                newByteArr.extend(hElse3Cnt)
                row += 1
                for i in range(else3Cnt):
                    railNo = ws.cell(row, 2).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)

                    cnt = ws.cell(row, 3).value
                    newByteArr.append(cnt)
                    for j in range(cnt):
                        pos = ws.cell(row, 4).value
                        newByteArr.append(pos)
                        railNo = ws.cell(row, 5).value
                        hRailNo = struct.pack("<h", railNo)
                        newByteArr.extend(hRailNo)
                        binIndex = ws.cell(row, 6).value
                        newByteArr.append(binIndex)

                        anime1 = ws.cell(row, 7).value
                        hAnime1 = struct.pack("<h", anime1)
                        newByteArr.extend(hAnime1)
                        anime2 = ws.cell(row, 8).value
                        hAnime2 = struct.pack("<h", anime2)
                        newByteArr.extend(hAnime2)
                        row += 1

                # 要素４
                ws = wb[tabList[9]]
                row = self.findLabel("else4", ws["A"])
                else4Cnt = ws.cell(row, 2).value
                hElse4Cnt = struct.pack("<h", else4Cnt)
                newByteArr.extend(hElse4Cnt)
                row += 1
                for i in range(else4Cnt):
                    railNo = ws.cell(row, 2).value
                    hRailNo = struct.pack("<h", railNo)
                    newByteArr.extend(hRailNo)
                    prevRailNo = ws.cell(row, 3).value
                    hPrevRailNo = struct.pack("<h", prevRailNo)
                    newByteArr.extend(hPrevRailNo)
                    for j in range(6):
                        tempF = struct.pack("<f", ws.cell(row, 4 + j).value)
                        newByteArr.extend(tempF)
                    row += 1

                # AMB情報
                dupNum = -1
                dupName = None
                ws = wb[tabList[10]]
                row = self.findLabel("AmbCnt", ws["A"])
                ambCnt = ws.cell(row, 2).value
                hAmbCnt = struct.pack("<h", ambCnt)
                newByteArr.extend(hAmbCnt)
                row = self.findLabel("index", ws["A"]) + 1
                for i in range(ambCnt):
                    newByteArr.append(ws.cell(row, 2).value)
                    fLength = struct.pack("<f", ws.cell(row, 3).value)
                    newByteArr.extend(fLength)
                    # RailNo, RailPos
                    for j in range(2):
                        tempH = struct.pack("<h", ws.cell(row, 4 + j).value)
                        newByteArr.extend(tempH)

                    # base pos_xyz base rot_xyz
                    for j in range(6):
                        tempF = struct.pack("<f", ws.cell(row, 6 + j).value)
                        newByteArr.extend(tempF)

                    for j in range(2):
                        newByteArr.append(ws.cell(row, 12 + j).value)

                    # mdl_no
                    mdl_no = ws.cell(row, 14).value
                    if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                        dupNum = i
                        dupName = ws.cell(row, 14).value
                    mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                    hMdlNo = struct.pack("<h", mdl_no)
                    newByteArr.extend(hMdlNo)

                    # pos xyz, dir xyz dir2 xyz
                    for j in range(9):
                        tempF = struct.pack("<f", ws.cell(row, 15 + j).value)
                        newByteArr.extend(tempF)

                    perF = struct.pack("<f", ws.cell(row, 24).value)
                    newByteArr.extend(perF)
                    row += 1

                    childCount = ws.cell(row, 13).value
                    newByteArr.append(childCount)
                    for j in range(childCount):
                        # mdl_no
                        mdl_no = ws.cell(row, 14).value
                        if self.isModelNameDup(mdl_no, smfNameList) and dupNum == -1:
                            dupNum = i
                            dupName = ws.cell(row, 14).value
                        mdl_no = self.getSmfModelIndex(i, mdl_no, smfNameList)
                        hMdlNo = struct.pack("<h", mdl_no)
                        newByteArr.extend(hMdlNo)

                        # pos xyz, dir xyz dir2 xyz
                        for k in range(9):
                            tempF = struct.pack("<f", ws.cell(row, 15 + k).value)
                            newByteArr.extend(tempF)

                        perF = struct.pack("<f", ws.cell(row, 24).value)
                        newByteArr.extend(perF)
                        row += 1
                    if childCount == 0:
                        row += 1

                if dupNum != -1:
                    warnMsg = textSetting.textList["infoList"]["I116"].format(dupNum, dupName)
                    result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
                    if not result:
                        return

            warnMsg = textSetting.textList["infoList"]["I117"]
            result = mb.askokcancel(title=textSetting.textList["warning"], message=warnMsg, icon="warning")
            if not result:
                return

            newBinFile = os.path.join(self.decryptFile.directory, self.decryptFile.filename + ".BIN")
            w = open(newBinFile, "wb")
            w.write(newByteArr)
            w.close()
            mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I114"])
            self.reloadFunc()
        except Exception:
            w = codecs.open("error.log", "w", "utf-8", "strict")
            w.write(traceback.format_exc())
            w.close()
            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])

    def findLabel(self, label, columns):
        for column in columns:
            if column.value == label:
                return column.row

    def isModelNameDup(self, modelName, smfNameList):
        if type(modelName) is str:
            modelNameList = [x[0] for x in smfNameList if x == modelName]
            if len(modelNameList) > 1:
                return True
            return False
        return False

    def getSmfModelName(self, modelIndex, smfNameList):
        if modelIndex < 0 or modelIndex >= len(smfNameList):
            return modelIndex
        modelName = smfNameList[modelIndex]
        if self.isModelNameDup(modelName, smfNameList):
            return modelIndex
        else:
            return modelName

    def getSmfModelIndex(self, i, modelValue, smfNameList):
        if type(modelValue) is str:
            if modelValue not in smfNameList:
                errorMsg = textSetting.textList["errorList"]["E96"].format(i, modelValue)
                mb.showerror(title=textSetting.textList["error"], message=errorMsg)
                return None
            modelIndex = smfNameList.index(modelValue)
            return modelIndex
        else:
            return modelValue
