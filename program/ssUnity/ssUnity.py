import os
import codecs
import tkinter
import openpyxl
from openpyxl.styles import PatternFill
import copy
import traceback
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import messagebox as mb
import program.textSetting as textSetting

from program.ssUnity.importPy.tkinterScrollbarTreeviewSSUnity import ScrollbarTreeviewSSUnity

unityFlag = True
try:
    from program.ssUnity.SSDecrypt.denDecrypt import DenDecrypt
    from program.ssUnity.SSDecrypt.resourcesDecrypt import ResourcesDecrypt
except ImportError:
    unityFlag = False

root = None
v_radio = None
v_fileName = None
v_select = None
v_search = None
searchEt = None
contentsLf = None
frame = None
monoCombo = None
extractBtn = None
loadAndSaveBtn = None
assetsSaveBtn = None
decryptFile = None

errorColorFill = PatternFill(patternType="solid", fgColor=textSetting.textList["excel"]["errorColor"])
warningColorFill = PatternFill(patternType="solid", fgColor=textSetting.textList["excel"]["warningColor"])


def deleteAllWidget():
    global contentsLf
    global v_btnList

    for children in contentsLf.winfo_children():
        children.destroy()
    
    for btn in v_btnList:
        btn["state"] = "disabled"


def createWidget(reloadFlag = False):
    global v_radio
    global v_select
    global v_btnList
    global v_search
    global searchEt
    global contentsLf
    global decryptFile
    global frame
    global monoCombo

    if not reloadFlag:
        v_search.set("")
        searchEt["state"] = "normal"

    frame = ScrollbarTreeviewSSUnity(contentsLf, v_select, v_btnList)
    col_tuple = (
        "treeNum",
        "treePathId",
        "treeName",
        "treeName2",
        "treeKind",
        "treeSize"
    )
    frame.tree["columns"] = col_tuple
    frame.tree.column("#0", width=0, stretch=False)

    frame.tree.column("treeNum", anchor=tkinter.CENTER, width=50, minwidth=50)
    frame.tree.column("treePathId", anchor=tkinter.CENTER)
    frame.tree.column("treeName", anchor=tkinter.CENTER)
    frame.tree.column("treeName2", anchor=tkinter.CENTER)
    frame.tree.column("treeKind", anchor=tkinter.CENTER)
    frame.tree.column("treeSize", anchor=tkinter.CENTER)
    frame.tree.heading("treeNum", text=textSetting.textList["ssUnity"]["headerNum"], anchor=tkinter.CENTER)
    frame.tree.heading("treePathId", text=textSetting.textList["ssUnity"]["headerPathId"], anchor=tkinter.CENTER)
    frame.tree.heading("treeName", text=textSetting.textList["ssUnity"]["headerName"], anchor=tkinter.CENTER)
    frame.tree.heading("treeName2", text=textSetting.textList["ssUnity"]["headerName"], anchor=tkinter.CENTER)
    frame.tree.heading("treeKind", text=textSetting.textList["ssUnity"]["headerKind"], anchor=tkinter.CENTER)
    frame.tree.heading("treeSize", text=textSetting.textList["ssUnity"]["headerSize"], anchor=tkinter.CENTER)

    if v_radio.get() == 0:
        for index, dataList in enumerate(decryptFile.allList):
            data = (index + 1, )
            data += ("", dataList[0], "", dataList[1], dataList[2])
            frame.tree.insert(parent="", index="end", iid=index, values=data)
        frame.tree["displaycolumns"] = (
            "treeNum",
            "treeName",
            "treeKind",
            "treeSize"
        )
    elif v_radio.get() == 1:
        if monoCombo["values"] == "":
            monoCombo["values"] = decryptFile.keyNameList
            monoCombo["state"] = "readonly"
            monoCombo.current(0)

        if monoCombo.current() == 0:
            for index, trainName in enumerate(decryptFile.trainNameList):
                data = (index + 1, )
                trainData = decryptFile.trainOrgInfoList[trainName]
                data += (trainData["num"], trainName, "", trainData["data"]["className"], trainData["data"]["size"])
                frame.tree.insert(parent="", index="end", iid=index, values=data)
            frame.tree["displaycolumns"] = (
                "treeNum",
                "treeName",
                "treeKind",
                "treeSize"
            )
        elif monoCombo.current() == 1:
            index = 0
            trainIdx = 0
            tagList = ["", "gray"]
            for trainModelName in decryptFile.trainModelNameList:
                changeMeshTexInfoList = decryptFile.changeMeshTexList[trainModelName]
                tags = tagList[trainIdx % 2]
                for changeMeshTexInfo in changeMeshTexInfoList:
                    data = (index + 1, )
                    meshTexInfo = changeMeshTexInfo["data"]["meshData"]
                    gameObjectName = changeMeshTexInfo["data"]["monoData"].m_GameObject.read().name
                    meshName = meshTexInfo[3]
                    data += (changeMeshTexInfo["num"], )
                    data += (trainModelName, "{0}({1})".format(gameObjectName, meshName), changeMeshTexInfo["data"]["className"], changeMeshTexInfo["data"]["size"])
                    frame.tree.insert(parent="", index="end", iid=index, values=data, tags=tags)
                    index += 1
                if len(changeMeshTexInfoList) > 0:
                    trainIdx += 1
            frame.tree.tag_configure("gray", background="#CCCCCC")
            frame.tree["displaycolumns"] = (
                "treeNum",
                "treePathId",
                "treeName",
                "treeName2",
                "treeKind",
                "treeSize"
            )


def selectGame():
    deleteAllWidget()
    changeButton()


def changeButton():
    global v_radio
    global v_fileName
    global v_select
    global v_search
    global searchEt
    global monoCombo
    global extractBtn
    global loadAndSaveBtn
    global assetsSaveBtn

    v_select.set("")
    v_fileName.set("")
    v_search.set("")
    searchEt["state"] = "readonly"
    if v_radio.get() == 0:
        extractBtn["text"] = textSetting.textList["ssUnity"]["extractFile"]
        extractBtn["command"] = extract
        extractBtn["state"] = "disabled"
        loadAndSaveBtn["text"] = textSetting.textList["ssUnity"]["saveFile"]
        loadAndSaveBtn["command"] = loadAndSave
        loadAndSaveBtn["state"] = "disabled"
        monoCombo.place_forget()
        assetsSaveBtn.place_forget()
    elif v_radio.get() == 1:
        monoCombo.set("")
        monoCombo["state"] = "disabled"
        monoCombo["values"] = ""
        monoCombo.place(relx=0.30, rely=0.03)
        extractBtn["text"] = textSetting.textList["ssUnity"]["extractCsv"]
        extractBtn["command"] = csvExtract
        extractBtn["state"] = "disabled"
        loadAndSaveBtn["text"] = textSetting.textList["ssUnity"]["saveCsv"]
        loadAndSaveBtn["command"] = csvLoadAndSave
        loadAndSaveBtn["state"] = "disabled"
        assetsSaveBtn.place(relx=0.78, rely=0.09)
        assetsSaveBtn["state"] = "disabled"


def changeResourceMonoEvent(event):
    changeResourceMono()


def changeResourceMono():
    global v_select
    global extractBtn
    global loadAndSaveBtn
    global assetsSaveBtn
    global frame

    v_select.set("")
    extractBtn["state"] = "disabled"
    loadAndSaveBtn["state"] = "disabled"
    assetsSaveBtn["state"] = "disabled"
    deleteAllWidget()
    createWidget()


def extract():
    global decryptFile
    global frame

    selectId = frame.tree.selection()[0]
    selectItem = frame.tree.set(selectId)
    num = int(selectItem["treeNum"]) - 1
    data = decryptFile.allList[num][-1]

    dataName = decryptFile.allList[num][0]
    excelFlag = False
    if dataName == "stagedata":
        excelFlag = True

    fileType = selectItem["treeKind"]
    if fileType == "TextAsset":
        ext = ".txt"
    elif fileType == "TextAsset(bytes)":
        ext = ".png"
    elif fileType == "AudioClip":
        ext = ".wav"
    filename = selectItem["treeName"]

    fileTypes = [(textSetting.textList["ssUnity"]["loadSaveFileLabel"], "*" + ext), ]
    defExt = ext
    if excelFlag:
        fileTypes.insert(0, (textSetting.textList["ssUnity"]["loadSaveFileExcelLabel"], "*.xlsx"))
        defExt = ".xlsx"

    file_path = fd.asksaveasfilename(initialfile=filename, filetypes=fileTypes, defaultextension=defExt)
    errorMsg = textSetting.textList["errorList"]["E89"]
    if file_path:
        try:
            if fileType == "AudioClip":
                for name, d in data.samples.items():
                    w = open(file_path, "wb")
                    w.write(d)
                    w.close()
            else:
                if excelFlag and os.path.splitext(file_path)[1].lower() == ".xlsx":
                    extractExcel(data.script.tobytes().decode(), file_path)
                else:
                    w = open(file_path, "wb")
                    w.write(data.script)
                    w.close()
                    mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I110"])
        except Exception:
            w = codecs.open("error.log", "w", "utf-8", "strict")
            w.write(traceback.format_exc())
            w.close()
            mb.showerror(title=textSetting.textList["error"], message=errorMsg)


def extractExcel(data, file_path):
    wb = openpyxl.Workbook()

    defSheetNameList = wb.sheetnames
    for sheetName in defSheetNameList:
        wb.remove(wb[sheetName])

    # TabList
    tabList = textSetting.textList["ssUnity"]["ssStageDataTabList"]

    errorLog = []
    warningLog = []
    mdlList = []
    for index, tabName in enumerate(tabList):
        wb.create_sheet(index=index, title=tabName)
        try:
            if not extractStageDataInfo(data, index, wb[tabName], mdlList, errorLog, warningLog):
                return
        except Exception:
            w = codecs.open("error.log", "w", "utf-8", "strict")
            w.write(traceback.format_exc())
            w.close()
            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
            return

    try:
        wb.save(file_path)
        mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I113"])
        if len(errorLog) > 0 or len(warningLog) > 0:
            dirPath = os.path.dirname(file_path)
            if len(errorLog) > 0:
                errPath = os.path.join(dirPath, "stageError.log")
                w = codecs.open(errPath, "w", "utf-8", "strict")
                for err in errorLog:
                    w.write(err + "\n")
                w.close()
            if len(warningLog) > 0:
                warnPath = os.path.join(dirPath, "stageWarning.log")
                w = codecs.open(warnPath, "w", "utf-8", "strict")
                for warn in warningLog:
                    w.write(warn + "\n")
                w.close()
            mb.showwarning(title=textSetting.textList["warning"], message=textSetting.textList["errorList"]["E118"])
    except PermissionError:
        mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E94"])


def extractStageDataInfo(data, sheetIndex, ws, mdlList, errorLog, warningLog):
    row = 1
    originDataList = data.split("\n")
    # コメント行を消す
    dataList = []
    for originData in originDataList:
        if originData.find("//") == 0:
            continue
        if not originData.strip():
            continue
        dataList.append(originData.strip())

    # ストーリー、配置情報
    if sheetIndex == 0:
        ws.cell(row, 1).value = "DEND_MAP_SS"
        row += 2

        search = "Story:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            ws.cell(row, 2).value = searchDataList[1]
        # Storyデータなし
        else:
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 2

        search = "Dir:"
        index = getSearchLine(dataList, search)
        if index != -1:
            searchDataList = getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    ws.cell(row, 2).value = int(searchDataList[1])
                except ValueError:
                    ws.cell(row, 2).value = searchDataList[1]
                    ws.cell(row, 2).fill = errorColorFill
                    errorLog.append(cntDataReadError(search, searchDataList[1]))
            # Dirデータなし
            else:
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(noDataError(search))
            row += 2

        search = "Track:"
        index = getSearchLine(dataList, search)
        if index != -1:
            searchDataList = getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    ws.cell(row, 2).value = int(searchDataList[1])
                except ValueError:
                    ws.cell(row, 2).value = searchDataList[1]
                    ws.cell(row, 2).fill = errorColorFill
                    errorLog.append(cntDataReadError(search, searchDataList[1]))
            # Trackデータなし
            else:
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(noDataError(search))
            row += 2

        search = "COMIC_DATA"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # COMIC_DATAの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            ws.cell(row, 1).value = searchDataList[0]
            # 「comic_」形式ではない
            if "comic_" not in searchDataList[0].lower():
                ws.cell(row, 1).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1, searchDataList[0]))
            row += 1
        row += 1

        search = "COMIC_IMAGE"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # COMIC_IMAGEの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            ws.cell(row, 1).value = searchDataList[0]
            # 「comic_img_」形式ではない
            if "comic_img_" not in searchDataList[0].lower():
                ws.cell(row, 1).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1, searchDataList[0]))
            row += 1
        row += 1

        search = "COMIC_SE"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # COMIC_SEの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            ws.cell(row, 1).value = searchDataList[0]
            # 「comic_se_」形式ではない
            if "comic_se_" not in searchDataList[0].lower():
                ws.cell(row, 1).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1, searchDataList[0]))
            row += 1
        row += 1

        search = "RailPos:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # RailPosの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # railPos(rail, bone)
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i + 1, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1))
                idx += 1
                colNum += 1

            # railPos p
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1))
            row += 1
        row += 1

        search = "FreeRun:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        row += 1

        for i in range(1):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # railPos(rail, bone)
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i + 1, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1))
                idx += 1
                colNum += 1

            # railPos p
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1))
            row += 1
        row += 1

        search = "VSPos:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # VSPosの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # railPos(rail, bone)
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i + 1, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1))
                idx += 1
                colNum += 1

            # railPos p
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1))
            row += 1
        row += 1

        search = "FadeImage:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # FadeImageの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # den, name
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    ws.cell(row, colNum).value = val
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1))
                idx += 1
                colNum += 1
            row += 1
        row += 1

    # 路線別画像データ
    elif sheetIndex == 1:
        search = "StageRes:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # StageResの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i + 1))
            idx += 1
            colNum += 1

            # den, name
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    ws.cell(row, colNum).value = val
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1))
                idx += 1
                colNum += 1
            row += 1
        row += 1

    # 画像設定情報
    elif sheetIndex == 2:
        search = "SetTexInfo:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # SetTexInfoの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # amb, amb_child, res_index
            for j in range(3):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # tex_type
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    tex_type = int(val)
                    ws.cell(row, colNum).value = tex_type
                except ValueError:
                    tex_type = -1
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                tex_type = -1
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # tex_index, change_index
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            if tex_type == 31:
                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(row, colNum).value = float(val)
                        except ValueError:
                            ws.cell(row, colNum).value = val
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i, val))
                    else:
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i))
                    idx += 1
                    colNum += 1
            row += 1
        row += 1
    # 駅名
    elif sheetIndex == 3:
        search = "STCnt:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # STCntの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # stIndex, rail
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # offset
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # name, jp, en
            for j in range(3):
                if realCnt > idx:
                    val = searchDataList[idx]
                    ws.cell(row, colNum).value = val
                    idx += 1
                    colNum += 1
                else:
                    break
            row += 1
        row += 1
    # ＣＰＵ切り替え
    elif sheetIndex == 4:
        search = "CPU:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # CPUの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # rail, train_no, run_type, min_len, max_len, max_speed, min_speed
            for j in range(7):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # p1(break)
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            row += 1
        row += 1

    # コミックスクリプト
    elif sheetIndex == 5:
        search = "ComicScript:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # ComicScriptの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # event_no, event_type, rail_no
            for j in range(3):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # offset
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            row += 1
        row += 1
    # 雨イベント
    elif sheetIndex == 6:
        search = "RainChecker:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # RainCheckerの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # event_no, event_type, rail_no
            for j in range(3):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # offset
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # param
            if realCnt > idx:
                paramCnt = len(searchDataList) - idx
                for j in range(paramCnt):
                    if searchDataList[idx + j].find("//") == 0:
                        break
                    val = searchDataList[idx + j]
                    try:
                        ws.cell(row, colNum + j).value = float(val)
                    except ValueError:
                        ws.cell(row, colNum + j).value = val
                        ws.cell(row, colNum + j).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
            row += 1
        row += 1
    # 土讃線スペシャル
    elif sheetIndex == 7:
        search = "DosanInfo:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # DosanInfoの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # event_no, event_type, rail_no
            for j in range(3):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # offset
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # param
            if realCnt > idx:
                paramCnt = len(searchDataList) - idx
                for j in range(paramCnt):
                    val = searchDataList[idx + j]
                    try:
                        ws.cell(row, colNum + j).value = float(val)
                    except ValueError:
                        ws.cell(row, colNum + j).value = val
                        ws.cell(row, colNum + j).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
            row += 1
        row += 1
    # モデル情報
    elif sheetIndex == 8:
        search = "MdlCnt:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # MdlCntの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # mdl_name
            if realCnt > idx:
                val = searchDataList[idx]
                ws.cell(row, colNum).value = val
                mdlList.append(val)
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # flg, flg, kasenchu_mdl
            for j in range(3):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1
            row += 1
        row += 1
    # レール情報
    elif sheetIndex == 9:
        search = "RailCnt:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # RailCntの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 2

        titleList = [
            "index",
            "prev_rail",
            "block",
            "pos_x",
            "pos_y",
            "pos_z",
            "dir_x",
            "dir_y",
            "dir_z",
            "mdl_no",
            "mdl_kasenchu",
            "per",
            "flg",
            "flg",
            "flg",
            "flg",
            "rail_data",
            "next_rail",
            "next_no",
            "prev_rail",
            "prev_no",
        ]

        for idx, title in enumerate(titleList):
            ws.cell(row, 1 + idx).value = title
            idx += 1
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # prev_rail, block
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # pos, dir (xyz)
            for j in range(6):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = float(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # mdl_no
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    mdl_no = int(val)
                    mdl_name = getModelName(mdl_no, mdlList)
                    ws.cell(row, colNum).value = mdl_name
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # mdl_kasenchu
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    mdl_kasenchu = int(val)
                    mdl_name = getModelName(mdl_kasenchu, mdlList)
                    ws.cell(row, colNum).value = mdl_name
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # per
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ws.cell(row, colNum).value = float(val)
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # flg, flg, flg, flg
            for j in range(4):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        flg = int(val)
                        ws.cell(row, colNum).value = toHex(flg)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # rail_data
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    rail_data = int(val)
                    ws.cell(row, colNum).value = rail_data
                except ValueError:
                    rail_data = 0
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
            else:
                rail_data = 0
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            for j in range(rail_data):
                for k in range(4):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(row, colNum).value = val
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i, val))
                    else:
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i))
                    idx += 1
                    colNum += 1
            row += 1
        row += 1
    # Pri情報
    elif sheetIndex == 10:
        search = "RailPri:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # RailPriの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i + 1, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i + 1))
                idx += 1
                colNum += 1
            row += 1
        row += 1

        search = "BtlPri:"
        index = getSearchLine(dataList, search)
        if index != -1:
            searchDataList = getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(row, 2).value = searchDataList[1]
                    ws.cell(row, 2).fill = errorColorFill
                    errorLog.append(cntDataReadError(search, searchDataList[1]))
            # BtlPriの数なし
            else:
                cnt = 0
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(noDataError(search))
            row += 1

            for i in range(cnt):
                searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(row, colNum).value = val
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i + 1, val))
                    else:
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                row += 1
            row += 1

        search = "NoDriftRail:"
        index = getSearchLine(dataList, search)
        if index != -1:
            searchDataList = getSplitAndRemoveEmptyData(dataList[index])
            ws.cell(row, 1).value = searchDataList[0]
            if len(searchDataList) > 1:
                try:
                    cnt = int(searchDataList[1])
                    ws.cell(row, 2).value = cnt
                except ValueError:
                    cnt = 0
                    ws.cell(row, 2).value = searchDataList[1]
                    ws.cell(row, 2).fill = errorColorFill
                    errorLog.append(cntDataReadError(search, searchDataList[1]))
            # NoDriftRailの数なし
            else:
                cnt = 0
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(noDataError(search))
            row += 1

            for i in range(cnt):
                searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
                realCnt = len(searchDataList)
                idx = 0
                colNum = idx + 1

                for j in range(2):
                    if realCnt > idx:
                        val = searchDataList[idx]
                        try:
                            ws.cell(row, colNum).value = int(val)
                        except ValueError:
                            ws.cell(row, colNum).value = val
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i + 1, val))
                    else:
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i + 1))
                    idx += 1
                    colNum += 1
                row += 1
            row += 1

    # AMB情報
    elif sheetIndex == 11:
        search = "AmbCnt:"
        index = getSearchLine(dataList, search)
        if index == -1:
            return failSearchError(search)
        searchDataList = getSplitAndRemoveEmptyData(dataList[index])
        ws.cell(row, 1).value = searchDataList[0]
        if len(searchDataList) > 1:
            try:
                cnt = int(searchDataList[1])
                ws.cell(row, 2).value = cnt
            except ValueError:
                cnt = 0
                ws.cell(row, 2).value = searchDataList[1]
                ws.cell(row, 2).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[1]))
        # AmbCntの数なし
        else:
            cnt = 0
            ws.cell(row, 2).fill = errorColorFill
            errorLog.append(noDataError(search))

        if len(searchDataList) > 2:
            try:
                sizeFlag = int(searchDataList[2])
                ws.cell(row, 3).value = sizeFlag
            except ValueError:
                sizeFlag = 0
                ws.cell(row, 3).value = searchDataList[2]
                ws.cell(row, 3).fill = errorColorFill
                errorLog.append(cntDataReadError(search, searchDataList[2]))
        else:
            sizeFlag = 0
        row += 2

        titleList = [
            "index",
            "rail",
            "length",
            "amd_data",
            "mdl_no",
            "parentIndex",
            "pos_x",
            "pos_y",
            "pos_z",
            "dir_x",
            "dir_y",
            "dir_z",
            "joint_dir_x",
            "joint_dir_y",
            "joint_dir_z",
            "per",
            "kasenchu_per",
        ]

        for idx, title in enumerate(titleList):
            ws.cell(row, 1 + idx).value = title
            idx += 1
        row += 1

        for i in range(cnt):
            searchDataList = getSplitAndRemoveEmptyData(dataList[index + i + 1])
            realCnt = len(searchDataList)
            idx = 0
            colNum = idx + 1

            # index
            if realCnt > idx:
                val = i
                ws.cell(row, colNum).value = val
            else:
                ws.cell(row, colNum).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
            idx += 1
            colNum += 1

            # rail_no, length
            for j in range(2):
                if realCnt > idx:
                    val = searchDataList[idx]
                    try:
                        ws.cell(row, colNum).value = int(val)
                    except ValueError:
                        ws.cell(row, colNum).value = val
                        ws.cell(row, colNum).fill = errorColorFill
                        errorLog.append(dataReadError(search, i, val))
                else:
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i))
                idx += 1
                colNum += 1

            # amb_data
            if realCnt > idx:
                val = searchDataList[idx]
                try:
                    ambData = int(val)
                    ws.cell(row, colNum).value = ambData
                except ValueError:
                    ws.cell(row, colNum).value = val
                    ws.cell(row, colNum).fill = errorColorFill
                    errorLog.append(dataReadError(search, i, val))
                    continue
            else:
                ambData = None
            idx += 1
            colNum += 1

            if ambData is None:
                ws.cell(row, 1).fill = errorColorFill
                ws.cell(row, 2).fill = errorColorFill
                ws.cell(row, 3).fill = errorColorFill
                ws.cell(row, 4).fill = errorColorFill
                errorLog.append(dataReadError(search, i))
                row += 1
            else:
                if ambData <= 0:
                    ws.cell(row, 1).fill = warningColorFill
                    ws.cell(row, 2).fill = warningColorFill
                    ws.cell(row, 3).fill = warningColorFill
                    ws.cell(row, 4).fill = warningColorFill
                    warningLog.append(ambDataWarning(i))
                    # ambDataは0個なのに、データがある
                    if len(searchDataList) > idx:
                        for j in range(idx, len(searchDataList)):
                            try:
                                ws.cell(row, j + 1).value = float(searchDataList[j])
                                ws.cell(row, j + 1).fill = warningColorFill
                            except ValueError:
                                ws.cell(row, j + 1).value = searchDataList[j]
                                ws.cell(row, j + 1).fill = errorColorFill
                    row += 1
                else:
                    for j in range(ambData):
                        colNum = 5
                        # mdl_no
                        if realCnt > idx:
                            val = searchDataList[idx]
                            try:
                                mdl_no = int(val)
                                mdl_name = getModelName(mdl_no, mdlList)
                                ws.cell(row, colNum).value = mdl_name
                            except ValueError:
                                ws.cell(row, colNum).value = val
                                ws.cell(row, colNum).fill = errorColorFill
                                errorLog.append(dataReadError(search, i, val))
                        else:
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i))
                        idx += 1
                        colNum += 1

                        # parentindex
                        if realCnt > idx:
                            val = searchDataList[idx]
                            try:
                                ws.cell(row, colNum).value = int(val)
                            except ValueError:
                                ws.cell(row, colNum).value = val
                                ws.cell(row, colNum).fill = errorColorFill
                                errorLog.append(dataReadError(search, i, val))
                        else:
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i))
                        idx += 1
                        colNum += 1

                        # pos, dir, joint_dir (xyz)
                        for j in range(9):
                            if realCnt > idx:
                                val = searchDataList[idx]
                                try:
                                    ws.cell(row, colNum).value = float(val)
                                except ValueError:
                                    ws.cell(row, colNum).value = val
                                    ws.cell(row, colNum).fill = errorColorFill
                                    errorLog.append(dataReadError(search, i, val))
                            else:
                                ws.cell(row, colNum).fill = errorColorFill
                                errorLog.append(dataReadError(search, i))
                            idx += 1
                            colNum += 1

                        # per
                        if realCnt > idx:
                            val = searchDataList[idx]
                            try:
                                ws.cell(row, colNum).value = float(val)
                            except ValueError:
                                ws.cell(row, colNum).value = val
                                ws.cell(row, colNum).fill = errorColorFill
                                errorLog.append(dataReadError(search, i, val))
                        else:
                            ws.cell(row, colNum).fill = errorColorFill
                            errorLog.append(dataReadError(search, i))
                        idx += 1
                        colNum += 1

                        # size_per
                        if sizeFlag >= 1:
                            if realCnt > idx:
                                val = searchDataList[idx]
                                try:
                                    ws.cell(row, colNum).value = float(val)
                                except ValueError:
                                    ws.cell(row, colNum).value = val
                                    ws.cell(row, colNum).fill = errorColorFill
                                    errorLog.append(dataReadError(search, i, val))
                            else:
                                ws.cell(row, colNum).fill = errorColorFill
                                errorLog.append(dataReadError(search, i))
                            idx += 1
                            colNum += 1
                        row += 1
                    moreFlag = False
                    # データがもっとある
                    while realCnt > idx + 12:
                        moreFlag = True
                        # mdl_no
                        val = searchDataList[idx]
                        try:
                            mdl_no = int(val)
                            mdl_name = getModelName(mdl_no, mdlList)
                            ws.cell(row - 1, colNum).value = mdl_name
                            ws.cell(row - 1, colNum).fill = warningColorFill
                        except ValueError:
                            ws.cell(row - 1, colNum).value = val
                            ws.cell(row - 1, colNum).fill = errorColorFill
                        idx += 1
                        colNum += 1

                        # parentindex
                        val = searchDataList[idx]
                        try:
                            ws.cell(row - 1, colNum).value = int(val)
                            ws.cell(row - 1, colNum).fill = warningColorFill
                        except ValueError:
                            ws.cell(row - 1, colNum).value = val
                            ws.cell(row - 1, colNum).fill = errorColorFill
                        idx += 1
                        colNum += 1

                        # pos, dir, joint_dir (xyz)
                        for j in range(9):
                            val = searchDataList[idx]
                            try:
                                ws.cell(row - 1, colNum).value = float(val)
                                ws.cell(row - 1, colNum).fill = warningColorFill
                            except ValueError:
                                ws.cell(row - 1, colNum).value = val
                                ws.cell(row - 1, colNum).fill = errorColorFill
                            idx += 1
                            colNum += 1

                        # per
                        val = searchDataList[idx]
                        try:
                            ws.cell(row - 1, colNum).value = float(val)
                            ws.cell(row - 1, colNum).fill = warningColorFill
                        except ValueError:
                            ws.cell(row - 1, colNum).value = val
                            ws.cell(row - 1, colNum).fill = errorColorFill
                        idx += 1
                        colNum += 1

                        # size_per
                        if sizeFlag >= 1:
                            if realCnt > idx:
                                val = searchDataList[idx]
                                try:
                                    ws.cell(row - 1, colNum).value = float(val)
                                    ws.cell(row - 1, colNum).fill = warningColorFill
                                except ValueError:
                                    ws.cell(row - 1, colNum).value = val
                                    ws.cell(row - 1, colNum).fill = errorColorFill
                                idx += 1
                                colNum += 1
                    if moreFlag:
                        warningLog.append(notUsedAmbData(i))
        row += 1
    return True


def noCntDataError(search):
    return textSetting.textList["errorList"]["E116"].format(search)


def cntDataReadError(search, data):
    return textSetting.textList["errorList"]["E117"].format(search, data)


def dataReadError(search, i, data="（データなし）"):
    return textSetting.textList["errorList"]["E114"].format(search, i, data)


def ambDataWarning(i):
    return textSetting.textList["errorList"]["E113"].format(i)


def notUsedAmbData(i):
    return textSetting.textList["errorList"]["E119"].format(i)


def outOfRangeError(search, cnt, i):
    mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E112"].format(search, cnt, i))
    return False


def failSearchError(search):
    mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E106"].format(search))
    return False


def failExcelSearchError(search, errMsgObj={}):
    errMsgObj["message"] = textSetting.textList["errorList"]["E107"].format(search)
    return False


def failExcelValueError(search, coord=None, errMsgObj={}):
    errMsgObj["message"] = textSetting.textList["errorList"]["E108"].format(search)
    if coord is not None:
        errMsgObj["message"] = textSetting.textList["errorList"]["E111"].format(search, coord)
    return False


def getSearchLine(dataList, search):
    for idx, data in enumerate(dataList):
        if data.find(search) == 0:
            return idx
    return -1


def getSearchLineCntEnd(dataList, search, count):
    searchIdx = -1
    for idx, data in enumerate(dataList):
        if data.find(search) == 0:
            idx = searchIdx
            break
    if searchIdx == -1:
        return searchIdx
    cnt = 0
    endIdx = -1
    for i in range(searchIdx + 1, len(dataList), 1):
        if dataList[i].find("//") != 0:
            cnt += 1
            if cnt >= count:
                endIdx = i
                break
    return endIdx

def getSplitAndRemoveEmptyData(data):
    sList = data.strip().split("\t")
    sList = list(filter(None, sList))
    sList = [x for x in sList if x.find("//") != 0]
    return sList


def getModelName(mdl_no, mdlList):
    if mdl_no < 0 or mdl_no >= len(mdlList):
        return mdl_no
    mdlName = mdlList[mdl_no]
    dupList = [x for x in mdlList if x == mdlName]
    if len(dupList) > 1:
        return mdl_no
    else:
        return mdlName


def getModelIndex(val, idx, mdlList, errMsgObj, search):
    if type(val) is str:
        if val not in mdlList:
            errMsgObj["message"] = textSetting.textList["errorList"]["E109"].format(val)
            return None
        tempList = [x for x in mdlList if x == val]
        if len(tempList) > 1:
            if search == "RailCnt:":
                errMsgObj["warning"] = textSetting.textList["infoList"]["I115"].format(idx, val)
            elif search == "AmbCnt:":
                errMsgObj["warning"] = textSetting.textList["infoList"]["I116"].format(idx, val)
        modelIndex = mdlList.index(val)
        return modelIndex
    else:
        return val


def toHex(num):
    return "0x{:02x}".format(num)


def loadAndSave():
    global decryptFile
    global frame

    selectId = frame.tree.selection()[0]
    selectItem = frame.tree.set(selectId)
    num = int(selectItem["treeNum"]) - 1
    data = decryptFile.allList[num][-1]

    dataName = decryptFile.allList[num][0]
    excelFlag = False
    if dataName == "stagedata":
        excelFlag = True

    fileType = selectItem["treeKind"]
    if fileType == "TextAsset":
        ext = ".txt"
    elif fileType == "TextAsset(bytes)":
        ext = ".png"
    elif fileType == "AudioClip":
        errorMsg = textSetting.textList["errorList"]["E90"]
        mb.showerror(title=textSetting.textList["error"], message=errorMsg)
        return

    fileTypes = [(textSetting.textList["ssUnity"]["loadSaveFileLabel"], "*" + ext), ]
    defExt = ext
    if excelFlag:
        fileTypes.insert(0, (textSetting.textList["ssUnity"]["loadSaveFileExcelLabel"], "*.xlsx"))
        defExt = ".xlsx"

    file_path = fd.askopenfilename(filetypes=fileTypes, defaultextension=defExt)
    if not file_path:
        return
    if not excelFlag or (excelFlag and os.path.splitext(file_path)[1].lower() != ".xlsx"):
        result = mb.askquestion(title=textSetting.textList["confirm"], message=textSetting.textList["infoList"]["I50"], icon="warning")
        if result == "no":
            return

    try:
        if fileType != "AudioClip":
            if excelFlag and os.path.splitext(file_path)[1].lower() == ".xlsx":
                errMsgObj = {}
                newLinesObj = {}
                if not loadExcelAndMerge(file_path, data, newLinesObj, errMsgObj):
                    mb.showerror(title=textSetting.textList["error"], message=errMsgObj["message"])
                    return
                if "warning" in errMsgObj:
                    result = mb.askquestion(title=textSetting.textList["confirm"], message=errMsgObj["warning"], icon="warning")
                    if result == "no":
                        return
                result = mb.askquestion(title=textSetting.textList["confirm"], message=textSetting.textList["infoList"]["I50"], icon="warning")
                if result == "no":
                    return
                data.script = bytearray("\n".join(newLinesObj["data"]).encode("utf-8"))
            else:
                with open(file_path, "rb") as f:
                    data.script = f.read()
            data.save()
        with open(decryptFile.filePath, "wb") as w:
            w.write(decryptFile.env.file.save())
        mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I51"])
        reloadFile()
    except Exception:
        w = codecs.open("error.log", "w", "utf-8", "strict")
        w.write(traceback.format_exc())
        w.close()
        mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])


def loadExcelAndMerge(file_path, data, newLinesObj, errMsgObj):
    originData = data.script.tobytes().decode()
    newLines = copy.deepcopy(originData).split("\n")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    tabList = textSetting.textList["ssUnity"]["ssStageDataTabList"]

    for tabName in tabList:
        if tabName not in wb.sheetnames:
            errMsgObj["message"] = textSetting.textList["errorList"]["E95"].format(tabName)
            return False

    ret = True
    ret &= findSearchAndSetCnt("Story:", wb[tabList[0]], newLines, dataFlag=False, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("Dir:", wb[tabList[0]], newLines, dataFlag=False, requiredFlag=False, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("Track:", wb[tabList[0]], newLines, dataFlag=False, requiredFlag=False, otherSearchList=["Dir:", "Story:"], errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("COMIC_DATA", wb[tabList[0]], newLines, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("COMIC_IMAGE", wb[tabList[0]], newLines, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("COMIC_SE", wb[tabList[0]], newLines, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("RailPos:", wb[tabList[0]], newLines, optionalRead=0, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("FreeRun:", wb[tabList[0]], newLines, headerDataFlag=False, optionalRead=0, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("VSPos:", wb[tabList[0]], newLines, optionalRead=0, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("FadeImage:", wb[tabList[0]], newLines, optionalRead=1, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("StageRes:", wb[tabList[1]], newLines, optionalRead=0, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("SetTexInfo:", wb[tabList[2]], newLines, optionalRead=2, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("STCnt:", wb[tabList[3]], newLines, optionalRead=3, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("CPU:", wb[tabList[4]], newLines, optionalRead=4, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("ComicScript:", wb[tabList[5]], newLines, optionalRead=5, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("RainChecker:", wb[tabList[6]], newLines, optionalRead=6, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("DosanInfo:", wb[tabList[7]], newLines, optionalRead=6, errMsgObj=errMsgObj)
    newMdlList = []
    ret &= findSearchAndSetCnt("MdlCnt:", wb[tabList[8]], newLines, optionalRead=5, mdlList=newMdlList, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("RailCnt:", wb[tabList[9]], newLines, optionalRead=7, mdlList=newMdlList, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("RailPri:", wb[tabList[10]], newLines, optionalRead=1, errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("BtlPri:", wb[tabList[10]], newLines, requiredFlag=False, optionalRead=1, otherSearchList=["RailPri:"], errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("NoDriftRail:", wb[tabList[10]], newLines, requiredFlag=False, optionalRead=1, otherSearchList=["BtlPri:", "RailPri:"], errMsgObj=errMsgObj)
    ret &= findSearchAndSetCnt("AmbCnt:", wb[tabList[11]], newLines, optionalRead=8, mdlList=newMdlList, errMsgObj=errMsgObj)

    newLinesObj["data"] = newLines
    return ret


def findLabel(search, columns):
    for column in columns:
        if column.value == search:
            return column.row
    return -1


def findSearchAndSetCnt(search, ws, newLines, headerDataFlag=True, dataFlag=True, requiredFlag=True, optionalRead=-1, otherSearchList=["Story:"], mdlList=[], errMsgObj={}):
    eIndex = findLabel(search, ws["A"])
    # エクセルで見つけられない
    if eIndex == -1:
        # 必須項目の場合、エラー
        if requiredFlag:
            return failExcelSearchError(search, errMsgObj=errMsgObj)
        # 必須項目ではない場合
        else:
            index = getSearchLine(newLines, search)
            # denで見つけた場合、消す
            if index != -1:
                if dataFlag:
                    delcnt = int(getSplitAndRemoveEmptyData(newLines[index])[1])
                else:
                    delcnt = 1
                index -= 1
                while newLines[index].find("//") == 0:
                    newLines.pop(index)
                    index -= 1
                index += 1
                newLines.pop(index)
                for i in range(delcnt):
                    newLines.pop(index)
            return True
    valList = []
    for cell in ws[eIndex]:
        if cell.value is None:
            break
        valList.append(str(cell.value))
    # データ数が必要な項目で、2個より少ない場合
    if headerDataFlag:
        if len(valList) < 2:
            return failExcelValueError(search, errMsgObj=errMsgObj)
        newLine = "\t".join(valList)
        try:
            newCnt = int(valList[1])
        except ValueError:
            newCnt = 1
    # データ数が必要ない、または1個のみ
    else:
        newLine = "".join(valList)
        newCnt = 1
    newLine += "\r"

    index = getSearchLine(newLines, search)
    if index == -1:
        # 必須項目の場合、エラー
        if requiredFlag:
            return failSearchError(search)
        # 必須項目ではないが、元のデータにない場合、追加
        else:
            for otherSearch in otherSearchList:
                index = getSearchLine(newLines, otherSearch)
                if index != -1:
                    break
            if dataFlag:
                cnt = int(getSplitAndRemoveEmptyData(newLines[index])[1])
                index += cnt
            index += 1
            newLines.insert(index, newLine)
            newLines.insert(index, "\r")

    if headerDataFlag:
        index = getSearchLine(newLines, search)
        try:
            originCnt = int(getSplitAndRemoveEmptyData(newLines[index])[1])
        except ValueError:
            originCnt = 1
    else:
        originCnt = 1
    newLines[index] = newLine

    # データ数通り、読み込む
    if dataFlag:
        newDataList = []

        startIndex = eIndex + 1
        if newCnt > 0:
            while True:
                val = ws.cell(startIndex, 1).value
                if val is not None and str(val).find("//") != 0:
                    break
                startIndex += 1

            # レールデータやAMB
            if optionalRead == 7 or optionalRead == 8:
                val = ws.cell(startIndex, 1).value
                if "index" in str(val):
                    startIndex += 1

        for i in range(newCnt):
            # デフォルト読み
            if optionalRead == -1:
                val = ws.cell(startIndex + i, 1).value
                if val is None:
                    coordinate = ws.cell(startIndex + i, 1).coordinate
                    return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                newDataList.append("{0}\r".format(val))
            # データ３つよみ
            elif optionalRead == 0:
                columnList = []
                for j in range(3):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # データ２つ読み
            elif optionalRead == 1:
                columnList = []
                for j in range(2):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # SetTexInfoの読み
            elif optionalRead == 2:
                columnList = []
                tex_type = -1
                for j in range(7):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                    if j == 4:
                        tex_type = int(val)
                val8 = ws.cell(startIndex + i, 8).value
                val9 = ws.cell(startIndex + i, 9).value
                if tex_type == 31 and val8 is not None and val9 is not None:
                    columnList.append("{0}".format(val8))
                    columnList.append("{0}".format(val9))
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # 駅名の読み
            elif optionalRead == 3:
                columnList = []
                for j in range(4):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                val5 = ws.cell(startIndex + i, 5).value
                val6 = ws.cell(startIndex + i, 6).value
                val7 = ws.cell(startIndex + i, 7).value
                if val5 is not None:
                    columnList.append("{0}".format(val5))
                if val6 is not None:
                    columnList.append("{0}".format(val6))
                if val7 is not None:
                    columnList.append("{0}".format(val7))
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # ＣＰＵの読み
            elif optionalRead == 4:
                columnList = []
                for j in range(8):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                val9 = ws.cell(startIndex + i, 9).value
                if val9 is not None:
                    columnList.append("{0}".format(val9))
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # データ５つの読み
            elif optionalRead == 5:
                columnList = []
                for j in range(5):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                    if j == 1 and search == "MdlCnt:":
                        mdlList.append(val)
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # イベントの読み
            elif optionalRead == 6:
                columnList = []
                for j in range(5):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                colIdx = 6
                while True:
                    val = ws.cell(startIndex + i, colIdx).value
                    if val is None:
                        break
                    columnList.append("{0}".format(val))
                    colIdx += 1
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # レールデータの読み
            elif optionalRead == 7:
                columnList = []
                for j in range(16):
                    val = ws.cell(startIndex + i, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex + i, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    # mdl_no
                    if j == 9:
                        val = getModelIndex(val, startIndex + i, mdlList, errMsgObj, search)
                        if val is None:
                            return False
                    # mdl_kasenchu
                    elif j == 10:
                        val = getModelIndex(val, startIndex + i, mdlList, errMsgObj, search)
                        if val is None:
                            return False
                    # flg
                    elif j >= 12 and j <= 15:
                        val = int(val, 16)
                    columnList.append("{0}".format(val))
                rail_data = ws.cell(startIndex + i, 17).value
                if rail_data is None:
                    coordinate = ws.cell(startIndex + i, 17).coordinate
                    return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                columnList.append("{0}".format(rail_data))

                for j in range(rail_data):
                    for k in range(4):
                        val = ws.cell(startIndex + i, 18 + 4*j + k).value
                        if val is None:
                            coordinate = ws.cell(startIndex + i, 18 + 4*j + k).coordinate
                            return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                        columnList.append("{0}".format(val))
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))
            # AMBデータの読み
            elif optionalRead == 8:
                columnList = []
                for j in range(3):
                    val = ws.cell(startIndex, 1 + j).value
                    if val is None:
                        coordinate = ws.cell(startIndex, 1 + j).coordinate
                        return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                    columnList.append("{0}".format(val))
                amb_data = ws.cell(startIndex, 4).value
                columnList.append("{0}".format(amb_data))

                if amb_data <= 0:
                    startIndex += 1
                else:
                    for j in range(amb_data):
                        for k in range(13):
                            val = ws.cell(startIndex, 5 + k).value
                            if val is None:
                                coordinate = ws.cell(startIndex, 5 + k).coordinate
                                return failExcelValueError(search, coordinate, errMsgObj=errMsgObj)
                            # mdl_no
                            if k == 0:
                                val = getModelIndex(val, startIndex, mdlList, errMsgObj, search)
                                if val is None:
                                    return False
                            columnList.append("{0}".format(val))
                        startIndex += 1
                columnLine = "\t".join(columnList)
                newDataList.append("{0}\r".format(columnLine))

        newDataList.reverse()

        index += 1
        delCnt = 0
        while delCnt < originCnt:
            if not newLines[index].strip():
                delCnt -= 1
            newLines.pop(index)
            delCnt += 1
        for i in range(newCnt):
            newLines.insert(index, newDataList[i])
    return True


def csvExtract():
    global monoCombo
    global decryptFile
    global frame

    selectId = frame.tree.selection()[0]
    selectItem = frame.tree.set(selectId)
    if monoCombo.current() == 0:
        trainName = selectItem["treeName"]
        data = decryptFile.trainOrgInfoList[trainName]["data"]["monoData"]
        filename = selectItem["treeName"] + ".csv"
        file_path = fd.asksaveasfilename(initialfile=filename, defaultextension="csv", filetypes=[("trainOrgInfo", "*.csv")])
        errorMsg = textSetting.textList["errorList"]["E7"]
        if file_path:
            try:
                w = codecs.open(file_path, "w", "utf-8-sig", "strict")
                trainOrgInfo = decryptFile.getTrainOrgInfo(data.raw_data)
                if trainOrgInfo is None:
                    decryptFile.printError()
                    errorMsg = textSetting.textList["errorList"]["E14"]
                    mb.showerror(title=textSetting.textList["error"], message=errorMsg)
                    return
                w.write("{0},{1}\n".format(textSetting.textList["ssUnity"]["csvNotchNum"], trainOrgInfo[1]))
                w.write("{0},{1}\n".format(textSetting.textList["ssUnity"]["csvOrgNum"], trainOrgInfo[5]))
                w.write("{0}\n".format(textSetting.textList["ssUnity"]["csvBodyClass"]))
                w.write(",".join(trainOrgInfo[6]))
                w.write("\n")
                w.write("{0}\n".format(textSetting.textList["ssUnity"]["csvBodyModel"]))
                w.write(",".join(trainOrgInfo[7]))
                w.write("\n")
                w.write("{0}\n".format(textSetting.textList["ssUnity"]["csvPantaModel"]))
                w.write(",".join(trainOrgInfo[8]))
                w.write("\n")
                w.write("{0}\n".format(textSetting.textList["ssUnity"]["csvBodyClassIndexList"]))
                w.write(",".join([str(x) for x in trainOrgInfo[9]]))
                w.write("\n")
                w.write("{0}\n".format(textSetting.textList["ssUnity"]["csvBodyModelIndexList"]))
                w.write(",".join([str(x) for x in trainOrgInfo[10]]))
                w.write("\n")
                w.write("{0}\n".format(textSetting.textList["ssUnity"]["csvPantaModelIndexList"]))
                w.write(",".join([str(x) for x in trainOrgInfo[11]]))
                w.close()
                mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I10"])
            except Exception:
                w = codecs.open("error.log", "a", "utf-8", "strict")
                w.write(traceback.format_exc())
                w.close()
                mb.showerror(title=textSetting.textList["error"], message=errorMsg)
    elif monoCombo.current() == 1:
        pathId = int(selectItem["treePathId"])
        trainModelName = selectItem["treeName"]
        meshTexTitleList = [
            "デフォルト",
            "回送",
            "試運転",
            "梅田",
            # 未使用
            "未詳1",
            "京橋",
            "名張",
            "難波",
            "三宮",
            "三田",
            "未詳2",
            "梅田",
            "品川",
            "大阪難波",
            "豊橋",
            "岐阜",
            "浅草",
            "日光",
            "池袋",
            # 未使用
            "渋谷(東急東横線)",
            "元町・中華街",
            "渋谷(東急田園都市線)",
            "中央林間",
        ]
        filename = selectItem["treeName"] + "_" + selectItem["treePathId"] + ".csv"
        file_path = fd.asksaveasfilename(initialfile=filename, defaultextension="csv", filetypes=[("changeMeshTex", "*.csv")])
        errorMsg = textSetting.textList["errorList"]["E7"]
        if file_path:
            try:
                w = codecs.open(file_path, "w", "utf-8-sig", "strict")
                meshTexInfoList = decryptFile.changeMeshTexList[trainModelName]
                meshTexInfo = [item for item in meshTexInfoList if item["num"] == pathId][0]
                for index, meshTexTitle in enumerate(meshTexTitleList):
                    w.write("{0},{1}\n".format(meshTexTitle, meshTexInfo["data"]["meshData"][-1][index]))
                w.close()
                mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I10"])
            except Exception:
                w = codecs.open("error.log", "a", "utf-8", "strict")
                w.write(traceback.format_exc())
                w.close()
                mb.showerror(title=textSetting.textList["error"], message=errorMsg)


def csvLoadAndSave():
    global monoCombo
    global decryptFile
    global frame

    selectId = frame.tree.selection()[0]
    selectItem = frame.tree.set(selectId)
    if monoCombo.current() == 0:
        trainName = selectItem["treeName"]
        file_path = fd.askopenfilename(defaultextension="csv", filetypes=[("trainOrgInfo", "*.csv")])
        if not file_path:
            return
        csvLines = None
        try:
            try:
                f = codecs.open(file_path, "r", "utf-8-sig", "strict")
                csvLines = f.readlines()
                f.close()
            except UnicodeDecodeError:
                f = codecs.open(file_path, "r", "shift-jis", "strict")
                csvLines = f.readlines()
                f.close()
            if not decryptFile.checkCsv(csvLines):
                decryptFile.printError()
                mb.showerror(title=textSetting.textList["error"], message=decryptFile.error)
            if not decryptFile.saveCsv(trainName):
                decryptFile.printError()
                mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
                return
            mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I111"])
            frame.tree.set(selectId, column="treeSize", value=str(decryptFile.trainOrgInfoList[trainName]["data"]["size"]))
        except Exception:
            w = codecs.open("error.log", "a", "utf-8", "strict")
            w.write(traceback.format_exc())
            w.close()
            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
    elif monoCombo.current() == 1:
        trainModelName = selectItem["treeName"]
        pathId = int(selectItem["treePathId"])
        file_path = fd.askopenfilename(defaultextension="csv", filetypes=[("changeMeshTex", "*.csv")])
        if not file_path:
            return
        try:
            try:
                f = codecs.open(file_path, "r", "utf-8-sig", "strict")
                csvLines = f.readlines()
                f.close()
            except UnicodeDecodeError:
                f = codecs.open(file_path, "r", "shift-jis", "strict")
                csvLines = f.readlines()
                f.close()
            if not decryptFile.saveChangeMeshTex(csvLines, trainModelName, pathId):
                decryptFile.printError()
                mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
                return
            mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I111"])
            changeMeshTexFilterInfo = [item for item in decryptFile.changeMeshTexList[trainModelName] if item["num"] == pathId][0]
            frame.tree.set(selectId, column="treeSize", value=str(changeMeshTexFilterInfo["data"]["size"]))
        except Exception:
            w = codecs.open("error.log", "a", "utf-8", "strict")
            w.write(traceback.format_exc())
            w.close()
            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])


def assetsSave():
    global decryptFile
    global assetsSaveBtn

    assetsSaveBtn["state"] = "disabled"
    assetsSaveBtn.update()
    try:
        if not decryptFile.saveAssets():
            decryptFile.printError()
            mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
            return
        mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I112"])
        reloadFile()
    except Exception:
        w = codecs.open("error.log", "a", "utf-8", "strict")
        w.write(traceback.format_exc())
        w.close()
        mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
    assetsSaveBtn["state"] = "normal"


def reloadFile():
    global v_radio
    global v_select
    global v_search
    global decryptFile
    global frame

    errorMsg = textSetting.textList["errorList"]["E14"]
    if decryptFile.filePath:
        try:
            if not decryptFile.open():
                decryptFile.printError()
                mb.showerror(title=textSetting.textList["error"], message=errorMsg)
                return

            selectId = None
            if v_select.get() != "":
                selectId = int(v_select.get())

            deleteAllWidget()
            createWidget(True)

            if v_search.get() != "":
                filterData()

            if selectId is not None:
                findFlag = False
                for idx, itemId in enumerate(frame.tree.get_children()):
                    item = frame.tree.item(itemId)
                    num = item["values"][0]
                    if selectId == num:
                        frame.tree.selection_set(itemId)
                        findFlag = True
                        break

                if findFlag:
                    if idx - 3 < 0:
                        frame.tree.see(0)
                    else:
                        frame.tree.see(idx - 3)

        except Exception:
            w = codecs.open("error.log", "a", "utf-8", "strict")
            w.write(traceback.format_exc())
            w.close()
            mb.showerror(title=textSetting.textList["error"], message=errorMsg)


def openFile():
    global unityFlag
    global v_radio
    global v_fileName
    global decryptFile

    if not unityFlag:
        msg = textSetting.textList["errorList"]["E91"]
        mb.showerror(title=textSetting.textList["error"], message=msg)
        return

    errorMsg = textSetting.textList["errorList"]["E14"]
    if v_radio.get() == 0:
        file_path = fd.askopenfilename(filetypes=[("DEND_SS", "*.den")])
        if file_path:
            filename = os.path.basename(file_path)
            v_fileName.set(filename)
            del decryptFile
            decryptFile = DenDecrypt(file_path)
            if not decryptFile.open():
                decryptFile.printError()
                mb.showerror(title=textSetting.textList["error"], message=errorMsg)
                return
            deleteAllWidget()
            createWidget()
    elif v_radio.get() == 1:
        file_path = fd.askopenfilename(filetypes=[("resources.assets", "resources.assets")])
        if file_path:
            filename = os.path.basename(file_path)
            v_fileName.set(filename)
            del decryptFile
            decryptFile = ResourcesDecrypt(file_path)
            if not decryptFile.open():
                decryptFile.printError()
                mb.showerror(title=textSetting.textList["error"], message=errorMsg)
                return
            deleteAllWidget()
            createWidget()


def filterList(event):
    global v_select
    global frame

    v_select.set("")
    if len(frame.tree.selection()) > 0:
        frame.tree.selection_remove(frame.tree.selection())
    filterData()


def filterData():
    global v_radio
    global v_search
    global decryptFile
    global frame
    global monoCombo

    if v_radio.get() == 0:
        for i in range(len(decryptFile.allList)):
            frame.tree.reattach(i, "", tkinter.END)
    elif v_radio.get() == 1:
        if monoCombo.current() == 0:
            for i in range(len(decryptFile.trainNameList)):
                frame.tree.reattach(i, "", tkinter.END)
        elif monoCombo.current() == 1:
            index = 0
            for trainModelName in decryptFile.trainModelNameList:
                changeMeshTexInfoList = decryptFile.changeMeshTexList[trainModelName]
                for changeMeshTexInfo in changeMeshTexInfoList:
                    index += 1
            for i in range(index):
                frame.tree.reattach(i, "", tkinter.END)

    search = v_search.get()
    for i in frame.tree.get_children():
        item = frame.tree.item(i)
        name = item["values"][2]
        if search.upper() not in name.upper():
            if v_radio.get() == 1 and monoCombo.current() == 1:
                name2 = item["values"][3]
                if search.upper() in name2.upper():
                    continue
            frame.tree.detach(i)


def call_ssUnity(rootTk, programFrame):
    global unityFlag
    global root
    global v_radio
    global v_fileName
    global v_select
    global v_btnList
    global v_search
    global monoCombo
    global extractBtn
    global loadAndSaveBtn
    global assetsSaveBtn
    global searchEt
    global contentsLf

    if not unityFlag:
        msg = textSetting.textList["errorList"]["E91"]
        mb.showerror(title=textSetting.textList["error"], message=msg)
        return

    root = rootTk
    v_radio = tkinter.IntVar()
    v_radio.set(0)

    v_fileName = tkinter.StringVar()
    fileNameEt = ttk.Entry(programFrame, textvariable=v_fileName, font=textSetting.textList["font2"], width=23, state="readonly", justify="center")
    fileNameEt.place(relx=0.053, rely=0.03)

    selectLb = ttk.Label(programFrame, text=textSetting.textList["ssUnity"]["selectNum"], font=textSetting.textList["font2"])
    selectLb.place(relx=0.05, rely=0.09)

    v_select = tkinter.StringVar()
    selectEt = ttk.Entry(programFrame, textvariable=v_select, font=textSetting.textList["font2"], width=6, state="readonly", justify="center")
    selectEt.place(relx=0.22, rely=0.09)

    monoCombo = ttk.Combobox(programFrame, font=textSetting.textList["font2"], state="disabled")
    monoCombo.place(relx=0.30, rely=0.03)
    monoCombo.bind("<<ComboboxSelected>>", changeResourceMonoEvent)
    monoCombo.place_forget()

    denRb = tkinter.Radiobutton(programFrame, text=textSetting.textList["ssUnity"]["editDenFile"], command=selectGame, variable=v_radio, value=0)
    denRb.place(relx=0.60, rely=0.03)
    resourcesRb = tkinter.Radiobutton(programFrame, text=textSetting.textList["ssUnity"]["editResourcesAssets"], command=selectGame, variable=v_radio, value=1)
    resourcesRb.place(relx=0.78, rely=0.03)

    extractBtn = ttk.Button(programFrame, text=textSetting.textList["ssUnity"]["extractFileLabel"], width=25, state="disabled", command=extract)
    extractBtn.place(relx=0.42, rely=0.09)
    loadAndSaveBtn = ttk.Button(programFrame, text=textSetting.textList["ssUnity"]["saveFileLabel"], width=25, state="disabled", command=loadAndSave)
    loadAndSaveBtn.place(relx=0.60, rely=0.09)

    assetsSaveBtn = ttk.Button(programFrame, text=textSetting.textList["ssUnity"]["saveResourcesAssets"], width=25, state="disabled", command=assetsSave)
    assetsSaveBtn.place(relx=0.78, rely=0.09)
    assetsSaveBtn.place_forget()

    v_btnList = [
        extractBtn,
        loadAndSaveBtn,
        assetsSaveBtn
    ]

    searchLb = ttk.Label(programFrame, text=textSetting.textList["ssUnity"]["searchText"], font=textSetting.textList["font2"])
    searchLb.place(relx=0.05, rely=0.15)

    v_search = tkinter.StringVar()
    searchEt = ttk.Entry(programFrame, textvariable=v_search, font=textSetting.textList["font7"], width=23, state="readonly", justify="center")
    searchEt.place(relx=0.10, rely=0.15)
    searchEt.bind("<KeyRelease>", filterList)

    contentsLf = ttk.LabelFrame(programFrame, text=textSetting.textList["ssUnity"]["scriptLabel"])
    contentsLf.place(relx=0.03, rely=0.20, relwidth=0.95, relheight=0.77)
