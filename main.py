import os
import platform
import requests
import webbrowser
import datetime
import sys
import configparser
import traceback
import tkinter
from tkinter import ttk
from tkinter import messagebox as mb

try:
    import ctypes, _ctypes
except ImportError:
    pass

import program.textSetting as textSetting
import program.comicscript.comicscript as comicscriptProgram
import program.mdlBin.mdlBin as mdlBinProgram
import program.mdlinfo.mdlinfo as mdlinfoProgram
import program.orgInfoEditor.orgInfoEditor as orgInfoEditorProgram
import program.musicEditor.musicEditor as musicEditorProgram
import program.fvtMaker.fvtMaker as fvtMakerProgram
import program.railEditor.railEditor as railEditorProgram
import program.smf.smf as smfProgram
import program.ssUnity.ssUnity as ssUnityProgram
import program.rsRail.rsRail as rsRailProgram
import program.appearance.rootFrameWidget as rootFrameWidget
from program.errorLogClass import ErrorLogObj


root = None
style = None
config_ini_path = None
v_comicscriptCheck = None
v_frameCheck = None
v_meshCheck = None
v_XYZCheck = None
v_mtrlCheck = None
v_flagGlbMode = None
v_modelNameMode = None
v_flagHexMode = None
v_ambReadMode = None
v_prog = None
selectedProgram = None
maxMenubarLen = None
version = 0
onlineUpdateVer = 0
updateFlag = False
menubar = None
rootFrameAppearance = None
rootFrameBackgroundColor = None
rootDarkModeFlag = False
darkModeDllPath = None
darkModeDll = None
errObj = ErrorLogObj()


def errorLog(message):
    errObj.write(message)


def resource_path(relative_path):
    bundle_dir = getattr(sys, "_MEIPASS", os.path.join(os.path.abspath(os.path.dirname(__file__))))
    return os.path.join(bundle_dir, relative_path)


def dll_path(relative_path):
    bundle_dir = getattr(sys, "_MEIPASS", os.path.join(os.path.abspath(os.path.dirname(__file__)), "program", "appearance", "dllData"))
    return os.path.join(bundle_dir, relative_path)


def clearRootFrame():
    global root

    children = root.winfo_children()
    for idx, child in enumerate(children):
        # menu
        if idx == 0:
            continue
        child.destroy()


def callProgram(programName):
    global root
    global selectedProgram
    global config_ini_path
    global rootFrameAppearance

    clearRootFrame()
    selectedProgram = programName
    if selectedProgram == "orgInfoEditor":
        orgInfoEditorProgram.call_orgInfoEditor(root, rootFrameAppearance)
    elif selectedProgram == "mdlBin":
        mdlBinProgram.call_mdlBin(root, rootFrameAppearance)
    elif selectedProgram == "mdlinfo":
        mdlinfoProgram.call_mdlinfo(root, rootFrameAppearance)
    elif selectedProgram == "comicscript":
        comicscriptProgram.call_comicscript(root, rootFrameAppearance)
    elif selectedProgram == "musicEditor":
        musicEditorProgram.call_musicEditor(root, rootFrameAppearance)
    elif selectedProgram == "fvtMaker":
        fvtMakerProgram.call_fvtMaker(root, rootFrameAppearance)
    elif selectedProgram == "railEditor":
        railEditorProgram.call_railEditor(root, config_ini_path, rootFrameAppearance)
    elif selectedProgram == "smf":
        smfProgram.call_smf(root, config_ini_path, rootFrameAppearance)
    elif selectedProgram == "SSUnity":
        ssUnityProgram.call_ssUnity(root, config_ini_path)
    elif selectedProgram == "rsRail":
        rsRailProgram.call_rsRail(root, rootFrameAppearance)

    delete_OptionMenu()
    if selectedProgram == "comicscript":
        add_comicscriptOptionMenu()
    elif selectedProgram == "smf":
        add_smfWriteOptionMenu()
    elif selectedProgram == "SSUnity" or selectedProgram == "railEditor":
        add_xlsxWriteOptionMenu()


def loadFile():
    global v_comicscriptCheck
    global v_frameCheck
    global v_meshCheck
    global v_XYZCheck
    global v_mtrlCheck
    global selectedProgram

    if selectedProgram == "orgInfoEditor":
        orgInfoEditorProgram.openFile()
    elif selectedProgram == "mdlBin":
        mdlBinProgram.openFile()
    elif selectedProgram == "mdlinfo":
        mdlinfoProgram.openFile()
    elif selectedProgram == "comicscript":
        comicscriptProgram.openFile(v_comicscriptCheck.get())
    elif selectedProgram == "musicEditor":
        musicEditorProgram.openFile()
    elif selectedProgram == "fvtMaker":
        fvtMakerProgram.openFile()
    elif selectedProgram == "railEditor":
        railEditorProgram.openFile()
    elif selectedProgram == "smf":
        smfProgram.openFile(v_frameCheck.get(), v_meshCheck.get(), v_XYZCheck.get(), v_mtrlCheck.get())
    elif selectedProgram == "SSUnity":
        ssUnityProgram.openFile()
    elif selectedProgram == "rsRail":
        rsRailProgram.openFile()
    else:
        mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E1"])


def configCheckOption(section, options, defaultValue="0"):
    global config_ini_path

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    if not configRead.has_option(section, options):
        if not configRead.has_section(section):
            configRead.add_section(section)
        configRead.set(section, options, defaultValue)

        try:
            f = open(config_ini_path, "w", encoding="utf-8")
            configRead.write(f)
            f.close()
        except PermissionError:
            errorLog(traceback.format_exc())

        return True
    return False


def add_comicscriptOptionMenu():
    global config_ini_path
    global v_comicscriptCheck

    if not os.path.exists(config_ini_path):
        writeDefaultConfig()

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    readErrorFlag = False
    if configCheckOption("COMICSCRIPT_GAME", "mode"):
        readErrorFlag = True

    if readErrorFlag:
        configRead.read(config_ini_path, encoding="utf-8")

    if menubar.entryconfig(tkinter.END) == menubar.entryconfig(maxMenubarLen):
        v_comicscriptCheck = tkinter.IntVar()
        v_comicscriptCheck.set(int(configRead.get("COMICSCRIPT_GAME", "mode")))
        comicscriptOptionMenu = tkinter.Menu(menubar, tearoff=False)
        gameList = textSetting.textList["menu"]["comicscript"]["gameList"]
        for gidx, game in enumerate(gameList):
            comicscriptOptionMenu.add_radiobutton(label=game, value=gidx, variable=v_comicscriptCheck, command=writeComicscriptConfig)
        menubar.add_cascade(label=textSetting.textList["menu"]["comicscript"]["name"], menu=comicscriptOptionMenu)


def add_smfWriteOptionMenu():
    global config_ini_path
    global v_frameCheck
    global v_meshCheck
    global v_XYZCheck
    global v_mtrlCheck
    global v_flagGlbMode
    global menubar
    global maxMenubarLen

    if not os.path.exists(config_ini_path):
        writeDefaultConfig()

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    readErrorFlag = False
    if configCheckOption("SMF_FRAME", "mode"):
        readErrorFlag = True
    if configCheckOption("SMF_MESH", "mode"):
        readErrorFlag = True
    if configCheckOption("SMF_XYZ", "mode"):
        readErrorFlag = True
    if configCheckOption("SMF_MTRL", "mode"):
        readErrorFlag = True
    if configCheckOption("GLB_WRITE", "mode"):
        readErrorFlag = True

    if readErrorFlag:
        configRead.read(config_ini_path, encoding="utf-8")

    if menubar.entryconfig(tkinter.END) == menubar.entryconfig(maxMenubarLen):
        v_frameCheck = tkinter.IntVar()
        v_frameCheck.set(int(configRead.get("SMF_FRAME", "mode")))
        v_meshCheck = tkinter.IntVar()
        v_meshCheck.set(int(configRead.get("SMF_MESH", "mode")))
        v_XYZCheck = tkinter.IntVar()
        v_XYZCheck.set(int(configRead.get("SMF_XYZ", "mode")))
        v_mtrlCheck = tkinter.IntVar()
        v_mtrlCheck.set(int(configRead.get("SMF_MTRL", "mode")))
        v_flagGlbMode = tkinter.IntVar()
        v_flagGlbMode.set(int(configRead.get("GLB_WRITE", "mode")))
        smfWriteOptionMenu = tkinter.Menu(menubar, tearoff=False)
        smfWriteOptionMenu.add_checkbutton(label=textSetting.textList["menu"]["smf"]["write"]["opt1"], variable=v_frameCheck, command=writeSmfConfig)
        smfWriteOptionMenu.add_checkbutton(label=textSetting.textList["menu"]["smf"]["write"]["opt2"], variable=v_meshCheck, command=writeSmfConfig)
        smfWriteOptionMenu.add_checkbutton(label=textSetting.textList["menu"]["smf"]["write"]["opt3"], variable=v_XYZCheck, command=writeSmfConfig)
        smfWriteOptionMenu.add_checkbutton(label=textSetting.textList["menu"]["smf"]["write"]["opt4"], variable=v_mtrlCheck, command=writeSmfConfig)
        smfWriteOptionMenu.add_separator()
        smfWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["smf"]["glb"]["opt1"], variable=v_flagGlbMode, value=0, command=writeGlbWriteConfig)
        smfWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["smf"]["glb"]["opt2"], variable=v_flagGlbMode, value=1, command=writeGlbWriteConfig)
        menubar.add_cascade(label=textSetting.textList["menu"]["smf"]["name"], menu=smfWriteOptionMenu)


def add_xlsxWriteOptionMenu():
    global config_ini_path
    global v_modelNameMode
    global v_flagHexMode
    global v_ambReadMode
    global menubar
    global maxMenubarLen

    if not os.path.exists(config_ini_path):
        writeDefaultConfig()

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    readErrorFlag = False
    if configCheckOption("MODEL_NAME_MODE", "mode"):
        readErrorFlag = True
    if configCheckOption("FLAG_MODE", "mode"):
        readErrorFlag = True
    if configCheckOption("AMB_READ_MODE", "mode"):
        readErrorFlag = True

    if readErrorFlag:
        configRead.read(config_ini_path, encoding="utf-8")

    if menubar.entryconfig(tkinter.END) == menubar.entryconfig(maxMenubarLen):
        v_modelNameMode = tkinter.IntVar()
        v_modelNameMode.set(int(configRead.get("MODEL_NAME_MODE", "mode")))
        v_flagHexMode = tkinter.IntVar()
        v_flagHexMode.set(int(configRead.get("FLAG_MODE", "mode")))
        v_ambReadMode = tkinter.IntVar()
        v_ambReadMode.set(int(configRead.get("AMB_READ_MODE", "mode")))
        xlsxWriteOptionMenu = tkinter.Menu(menubar, tearoff=False)
        xlsxWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["SSUnity"]["write"]["model1"], variable=v_modelNameMode, value=0, command=writeXlsxConfig)
        xlsxWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["SSUnity"]["write"]["model2"], variable=v_modelNameMode, value=1, command=writeXlsxConfig)
        xlsxWriteOptionMenu.add_separator()
        xlsxWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["SSUnity"]["write"]["flag1"], variable=v_flagHexMode, value=0, command=writeXlsxConfig)
        xlsxWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["SSUnity"]["write"]["flag2"], variable=v_flagHexMode, value=1, command=writeXlsxConfig)
        xlsxWriteOptionMenu.add_separator()
        xlsxWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["SSUnity"]["write"]["ambRead1"], variable=v_ambReadMode, value=0, command=writeXlsxConfig)
        xlsxWriteOptionMenu.add_radiobutton(label=textSetting.textList["menu"]["SSUnity"]["write"]["ambRead2"], variable=v_ambReadMode, value=1, command=writeXlsxConfig)
        menubar.add_cascade(label=textSetting.textList["menu"]["SSUnity"]["name"], menu=xlsxWriteOptionMenu)


def delete_OptionMenu():
    global menubar
    global maxMenubarLen

    if menubar.index(tkinter.END) > maxMenubarLen:
        menubar.delete(maxMenubarLen + 1)


def writeDefaultConfig():
    global config_ini_path
    global version

    if platform.system() == "Windows":
        try:
            config_ini_folder = os.path.dirname(config_ini_path)
            if not os.path.exists(config_ini_folder):
                os.makedirs(config_ini_folder)

            config = configparser.RawConfigParser()
            config.add_section("COMICSCRIPT_GAME")
            config.set("COMICSCRIPT_GAME", "mode", 0)

            config.add_section("SMF_FRAME")
            config.set("SMF_FRAME", "mode", 0)
            config.add_section("SMF_MESH")
            config.set("SMF_MESH", "mode", 0)
            config.add_section("SMF_XYZ")
            config.set("SMF_XYZ", "mode", 0)
            config.add_section("SMF_MTRL")
            config.set("SMF_MTRL", "mode", 0)
            config.add_section("GLB_WRITE")
            config.set("GLB_WRITE", "mode", 0)

            config.add_section("MODEL_NAME_MODE")
            config.set("MODEL_NAME_MODE", "mode", 0)
            config.add_section("FLAG_MODE")
            config.set("FLAG_MODE", "mode", 0)
            config.add_section("AMB_READ_MODE")
            config.set("AMB_READ_MODE", "mode", 0)

            config.add_section("UPDATE")
            config.set("UPDATE", "time", "2000/01/01")

            f = open(config_ini_path, "w", encoding="utf-8")
            config.write(f)
            f.close()
        except PermissionError:
            errorLog(traceback.format_exc())


def writeComicscriptConfig():
    global v_comicscriptCheck
    global config_ini_path

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    configRead.set("COMICSCRIPT_GAME", "mode", str(v_comicscriptCheck.get()))

    try:
        f = open(config_ini_path, "w", encoding="utf-8")
        configRead.write(f)
        f.close()
    except PermissionError:
        errorLog(traceback.format_exc())


def writeSmfConfig():
    global v_frameCheck
    global v_meshCheck
    global v_XYZCheck
    global v_mtrlCheck
    global config_ini_path

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    configRead.set("SMF_FRAME", "mode", str(v_frameCheck.get()))
    configRead.set("SMF_MESH", "mode", str(v_meshCheck.get()))
    configRead.set("SMF_XYZ", "mode", str(v_XYZCheck.get()))
    configRead.set("SMF_MTRL", "mode", str(v_mtrlCheck.get()))

    try:
        f = open(config_ini_path, "w", encoding="utf-8")
        configRead.write(f)
        f.close()
    except PermissionError:
        errorLog(traceback.format_exc())


def writeGlbWriteConfig():
    global v_flagGlbMode
    global config_ini_path

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    configRead.set("GLB_WRITE", "mode", str(v_flagGlbMode.get()))

    try:
        f = open(config_ini_path, "w", encoding="utf-8")
        configRead.write(f)
        f.close()
    except PermissionError:
        errorLog(traceback.format_exc())


def writeXlsxConfig():
    global v_modelNameMode
    global v_flagHexMode
    global v_ambReadMode
    global config_ini_path

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")

    configRead.set("MODEL_NAME_MODE", "mode", str(v_modelNameMode.get()))
    configRead.set("FLAG_MODE", "mode", str(v_flagHexMode.get()))
    configRead.set("AMB_READ_MODE", "mode", str(v_ambReadMode.get()))

    try:
        f = open(config_ini_path, "w", encoding="utf-8")
        configRead.write(f)
        f.close()
    except PermissionError:
        errorLog(traceback.format_exc())


def getUpdateVer():
    global version
    global onlineUpdateVer
    global updateFlag

    path = resource_path("ver.txt")
    f = open(path, "r", encoding="utf-8")
    line = f.read()
    f.close()
    version = line.strip()

    try:
        url = "https://raw.githubusercontent.com/khttemp/dend-mod-gui/main/ver.txt"
        response = requests.get(url)
        if response.status_code == requests.codes.ok:
            onlineUpdateVer = response.text

            if version != onlineUpdateVer:
                configCheckOption("UPDATE", "time", "2000/01/01")

                configRead = configparser.ConfigParser()
                configRead.read(config_ini_path, encoding="utf-8")

                localDateStr = configRead.get("UPDATE", "time")
                localDate = datetime.datetime.strptime(localDateStr, "%Y/%m/%d").date()
                currentDate = datetime.datetime.now().date()
                if (localDate - currentDate).days < 0:
                    updateFlag = True
    except Exception:
        errorLog(traceback.format_exc())


def confirmUpdate():
    global onlineUpdateVer
    global updateFlag

    if updateFlag:
        msg = textSetting.textList["update"]["message"].format(onlineUpdateVer)
        result = mb.askyesno(title=textSetting.textList["update"]["title"], message=msg)
        if result == tkinter.YES:
            webbrowser.open_new("https://github.com/khttemp/dend-mod-gui/releases")

        try:
            configRead = configparser.ConfigParser()
            configRead.read(config_ini_path, encoding="utf-8")

            currentTime = datetime.datetime.now()
            currentDate = datetime.datetime.strftime(currentTime, "%Y/%m/%d")
            configRead.set("UPDATE", "time", currentDate)

            f = open(config_ini_path, "w", encoding="utf-8")
            configRead.write(f)
            f.close()
        except PermissionError:
            errorLog(traceback.format_exc())


def readRootFrameAppearance():
    global root
    global menubar
    global style
    global config_ini_path
    global rootFrameAppearance
    global rootFrameBackgroundColor
    global rootDarkModeFlag
    global darkModeDllPath
    global darkModeDll

    configCheckOption("ROOT_FRAME", "bg_color", "SystemButtonFace")
    configCheckOption("ROOT_FRAME", "dark_mode")
    configCheckOption("ROOT_FRAME", "theme", "vista")
    configCheckOption("LABEL", "fg_color", "SystemWindowText")
    configCheckOption("LABELFRAME_LABEL", "fg_color", "SystemWindowText")
    configCheckOption("RADIO", "fg_color", "SystemWindowText")
    configCheckOption("TREEVIEW", "bg_color", "SystemWindow")
    configCheckOption("TREEVIEW", "fg_color", "SystemWindowText")
    configCheckOption("TREEVIEW", "sel_bg_color", "SystemHighlight")
    configCheckOption("TREEVIEW", "sel_fg_color", "SystemWindow")
    configCheckOption("BUTTON", "fg_color", "SystemWindowText")
    configCheckOption("ENTRY", "fg_color", "SystemWindowText")
    configCheckOption("TREEVIEW", "field_bg_color", "SystemWindow")
    configCheckOption("TREEVIEW_HEADER", "bg_color", "SystemButtonFace")
    configCheckOption("TREEVIEW_HEADER", "fg_color", "SystemWindowText")
    configCheckOption("COMBOBOX", "bg_color", "SystemWindow")
    configCheckOption("COMBOBOX", "fg_color", "SystemWindowText")
    configCheckOption("COMBOBOX", "sel_bg_color", "SystemHighlight")
    configCheckOption("COMBOBOX", "sel_fg_color", "SystemWindow")
    configCheckOption("RADIO", "indicator_color", "SystemWindow")
    configCheckOption("RADIO", "sel_indicator_color", "SystemWindowText")

    configRead = configparser.ConfigParser()
    configRead.read(config_ini_path, encoding="utf-8")
    rootDarkModeFlag = int(configRead.get("ROOT_FRAME", "dark_mode")) > 0
    rootFrameBackgroundColor = configRead.get("ROOT_FRAME", "bg_color")
    root["bg"] = rootFrameBackgroundColor

    labelForegroundColor = configRead.get("LABEL", "fg_color")
    labelframeLabelForegroundColor = configRead.get("LABELFRAME_LABEL", "fg_color")
    radioForegroundColor = configRead.get("RADIO", "fg_color")

    treeviewBackgroundColor = configRead.get("TREEVIEW", "bg_color")
    treeviewForegroundColor = configRead.get("TREEVIEW", "fg_color")
    treeviewSelectedBackgroundColor = configRead.get("TREEVIEW", "sel_bg_color")
    treeviewSelectedForegroundColor = configRead.get("TREEVIEW", "sel_fg_color")

    buttonForegroundColor = configRead.get("BUTTON", "fg_color")
    entryForegroundColor = configRead.get("ENTRY", "fg_color")

    treeviewFieldBackgroundColor = configRead.get("TREEVIEW", "field_bg_color")
    treeviewHeaderBackgroundColor = configRead.get("TREEVIEW_HEADER", "bg_color")
    treeviewHeaderForegroundColor = configRead.get("TREEVIEW_HEADER", "fg_color")

    comboboxBackgroundColor = configRead.get("COMBOBOX", "bg_color")
    comboboxForegroundColor = configRead.get("COMBOBOX", "fg_color")
    comboboxSelectedBackgroundColor = configRead.get("COMBOBOX", "sel_bg_color")
    comboboxSelectedForegroundColor = configRead.get("COMBOBOX", "sel_fg_color")

    indicatorColor = configRead.get("RADIO", "indicator_color")
    indicatorSelectedColor = configRead.get("RADIO", "sel_indicator_color")

    if platform.system() == "Windows":
        try:
            if rootDarkModeFlag:
                darkModeDllPath = dll_path("tablacusdark64.dll")
                darkModeDll = ctypes.CDLL(darkModeDllPath)
                rootDarkModeFlag = True
        except Exception:
            rootDarkModeFlag = False
            errorLog(traceback.format_exc())

    themeName = configRead.get("ROOT_FRAME", "theme")
    style.theme_use(themeName)

    style.configure("custom.TLabel", background=rootFrameBackgroundColor, foreground=labelForegroundColor)
    style.configure("custom.red.TLabel", background=rootFrameBackgroundColor, foreground="red")
    style.configure("custom.blue.TLabel", background=rootFrameBackgroundColor, foreground="blue")
    style.configure("custom.green.TLabel", background=rootFrameBackgroundColor, foreground="green")
    style.configure("custom.444444.TLabel", background=rootFrameBackgroundColor, foreground="#444444")
    style.configure("custom.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor)
    style.configure("custom.update.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font7"], width=5, disabledbackground=rootFrameBackgroundColor)
    style.configure("custom.listbox.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font2"], width=5)
    style.configure("custom.paste.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font2"], width=10)
    style.configure("custom.elsePerf.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font7"])
    style.configure("custom.TRadiobutton", background=rootFrameBackgroundColor, foreground=radioForegroundColor)
    style.configure("custom.TCheckbutton", background=rootFrameBackgroundColor, foreground=radioForegroundColor, font=textSetting.textList["font2"])
    style.configure("custom.railFlag.TCheckbutton", background=rootFrameBackgroundColor, foreground=radioForegroundColor)
    style.configure("custom.TLabelframe", background=rootFrameBackgroundColor)
    style.configure("custom.TLabelframe.Label", background=rootFrameBackgroundColor, foreground=labelframeLabelForegroundColor)
    style.configure("custom.TFrame", background=rootFrameBackgroundColor)
    style.configure("custom.TSeparator", background=rootFrameBackgroundColor)
    style.configure("custom.Treeview", background=treeviewBackgroundColor, foreground=treeviewForegroundColor, fieldbackground=treeviewFieldBackgroundColor)
    style.configure("custom.Treeview.Heading", background=treeviewHeaderBackgroundColor, foreground=treeviewHeaderForegroundColor)
    style.configure("custom.TMenubutton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor)
    style.configure("custom.TSpinbox", fieldbackground=rootFrameBackgroundColor, foreground=buttonForegroundColor)
    style.map("custom.TRadiobutton", indicatorcolor=[("!selected", indicatorColor), ("selected", indicatorSelectedColor)])
    style.map("custom.TCheckbutton", indicatorcolor=[("!selected", indicatorColor), ("selected", indicatorSelectedColor)])
    style.map("custom.railFlag.TCheckbutton", indicatorcolor=[("!selected", indicatorColor), ("selected", indicatorSelectedColor)])
    style.map("custom.TEntry", background=[("!readonly", rootFrameBackgroundColor), ("readonly", rootFrameBackgroundColor)], fieldbackground=[("!readonly", rootFrameBackgroundColor), ("readonly", rootFrameBackgroundColor)], foreground=[("!readonly", entryForegroundColor), ("readonly", entryForegroundColor)])
    style.map("custom.Horizontal.TScrollbar", background=[("!disabled", rootFrameBackgroundColor), ("disabled", rootFrameBackgroundColor)])
    style.map("custom.Vertical.TScrollbar", background=[("!disabled", rootFrameBackgroundColor), ("disabled", rootFrameBackgroundColor)])
    style.map("custom.TCombobox", background=[("readonly", rootFrameBackgroundColor), ("disabled", rootFrameBackgroundColor)], fieldbackground=[("readonly", comboboxBackgroundColor), ("disabled", comboboxBackgroundColor)], foreground=[("readonly", comboboxForegroundColor), ("disabled", comboboxForegroundColor)])
    style.map("custom.Treeview", background=[("selected", treeviewSelectedBackgroundColor)], foreground=[("selected", treeviewSelectedForegroundColor)])
    root.option_add("*TCombobox*Listbox.background", comboboxBackgroundColor)
    root.option_add("*TCombobox*Listbox.foreground", comboboxForegroundColor)
    root.option_add("*TCombobox*Listbox.selectBackground", comboboxSelectedBackgroundColor)
    root.option_add("*TCombobox*Listbox.selectForeground", comboboxSelectedForegroundColor)

    rootFrameAppearance = rootFrameWidget.RootFrameAppearance(root, config_ini_path, labelForegroundColor, rootFrameBackgroundColor, configRead)


def editRootFrameAppearance():
    global root
    global style
    global v_prog
    global selectedProgram
    global config_ini_path
    global rootFrameAppearance
    global rootFrameBackgroundColor
    global rootDarkModeFlag
    global darkModeDllPath
    global darkModeDll

    rootFrameAppearance.editRootFrameAppearance()
    if rootFrameAppearance.reloadFlag:
        themeName = rootFrameAppearance.themeName
        newRootDarkModeFlag = rootFrameAppearance.darkModeFlag
        rootFrameBackgroundColor = rootFrameAppearance.bgColor
        root["bg"] = rootFrameBackgroundColor
        labelForegroundColor = rootFrameAppearance.labelFgColor
        labelframeLabelForegroundColor = rootFrameAppearance.labelframeFgColor
        radioForegroundColor = rootFrameAppearance.radioFgColor

        treeviewBackgroundColor = rootFrameAppearance.treeviewBgColor
        treeviewForegroundColor = rootFrameAppearance.treeviewFgColor
        treeviewSelectedBackgroundColor = rootFrameAppearance.treeviewSelBgColor
        treeviewSelectedForegroundColor = rootFrameAppearance.treeviewSelFgColor

        buttonForegroundColor = rootFrameAppearance.buttonFgColor
        entryForegroundColor = rootFrameAppearance.entryFgColor

        treeviewFieldBackgroundColor = rootFrameAppearance.treeviewFieldBgColor
        treeviewHeaderBackgroundColor = rootFrameAppearance.treeviewHeaderBgColor
        treeviewHeaderForegroundColor = rootFrameAppearance.treeviewHeaderFgColor

        comboboxBackgroundColor = rootFrameAppearance.comboboxBgColor
        comboboxForegroundColor = rootFrameAppearance.comboboxFgColor
        comboboxSelectedBackgroundColor = rootFrameAppearance.comboboxSelBgColor
        comboboxSelectedForegroundColor = rootFrameAppearance.comboboxSelFgColor

        indicatorColor = rootFrameAppearance.indicatorColor
        indicatorSelectedColor = rootFrameAppearance.indicatorSelColor

        if platform.system() == "Windows":
            try:
                if rootDarkModeFlag != newRootDarkModeFlag:
                    if newRootDarkModeFlag:
                        darkModeDll = ctypes.CDLL(darkModeDllPath)
                    else:
                        _ctypes.FreeLibrary(darkModeDll._handle)
                        del darkModeDll
                        darkModeDll = None
                    rootDarkModeFlag = newRootDarkModeFlag
            except Exception:
                errorLog(traceback.format_exc())

        style.theme_use(themeName)
        style.configure("custom.TLabel", background=rootFrameBackgroundColor, foreground=labelForegroundColor)
        style.configure("custom.red.TLabel", background=rootFrameBackgroundColor, foreground="red")
        style.configure("custom.blue.TLabel", background=rootFrameBackgroundColor, foreground="blue")
        style.configure("custom.green.TLabel", background=rootFrameBackgroundColor, foreground="green")
        style.configure("custom.444444.TLabel", background=rootFrameBackgroundColor, foreground="#444444")
        style.configure("custom.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor)
        style.configure("custom.update.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font7"], width=5, disabledbackground=rootFrameBackgroundColor)
        style.configure("custom.listbox.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font2"], width=5)
        style.configure("custom.paste.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font2"], width=10)
        style.configure("custom.elsePerf.TButton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor, font=textSetting.textList["font7"])
        style.configure("custom.TRadiobutton", background=rootFrameBackgroundColor, foreground=radioForegroundColor)
        style.configure("custom.TCheckbutton", background=rootFrameBackgroundColor, foreground=radioForegroundColor, font=textSetting.textList["font2"])
        style.configure("custom.railFlag.TCheckbutton", background=rootFrameBackgroundColor, foreground=radioForegroundColor)
        style.configure("custom.TLabelframe", background=rootFrameBackgroundColor)
        style.configure("custom.TLabelframe.Label", background=rootFrameBackgroundColor, foreground=labelframeLabelForegroundColor)
        style.configure("custom.TFrame", background=rootFrameBackgroundColor)
        style.configure("custom.TSeparator", background=rootFrameBackgroundColor)
        style.configure("custom.Treeview", background=treeviewBackgroundColor, foreground=treeviewForegroundColor, fieldbackground=treeviewFieldBackgroundColor)
        style.configure("custom.Treeview.Heading", background=treeviewHeaderBackgroundColor, foreground=treeviewHeaderForegroundColor)
        style.configure("custom.TMenubutton", background=rootFrameBackgroundColor, foreground=buttonForegroundColor)
        style.configure("custom.TSpinbox", fieldbackground=rootFrameBackgroundColor, foreground=buttonForegroundColor)
        style.map("custom.TRadiobutton", indicatorcolor=[("!selected", indicatorColor), ("selected", indicatorSelectedColor)])
        style.map("custom.TCheckbutton", indicatorcolor=[("!selected", indicatorColor), ("selected", indicatorSelectedColor)])
        style.map("custom.railFlag.TCheckbutton", indicatorcolor=[("!selected", indicatorColor), ("selected", indicatorSelectedColor)])
        style.map("custom.TEntry", background=[("!readonly", rootFrameBackgroundColor), ("readonly", rootFrameBackgroundColor)], fieldbackground=[("!readonly", rootFrameBackgroundColor), ("readonly", rootFrameBackgroundColor)], foreground=[("!readonly", entryForegroundColor), ("readonly", entryForegroundColor)])
        style.map("custom.Horizontal.TScrollbar", background=[("!disabled", rootFrameBackgroundColor), ("disabled", rootFrameBackgroundColor)])
        style.map("custom.Vertical.TScrollbar", background=[("!disabled", rootFrameBackgroundColor), ("disabled", rootFrameBackgroundColor)])
        style.map("custom.TCombobox", background=[("readonly", rootFrameBackgroundColor), ("disabled", rootFrameBackgroundColor)], fieldbackground=[("readonly", comboboxBackgroundColor), ("disabled", comboboxBackgroundColor)], foreground=[("readonly", comboboxForegroundColor), ("disabled", comboboxForegroundColor)])
        style.map("custom.Treeview", background=[("selected", treeviewSelectedBackgroundColor)], foreground=[("selected", treeviewSelectedForegroundColor)])
        mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["appearance"]["success"])


def guiMain():
    global root
    global style
    global config_ini_path
    global v_frameCheck
    global v_meshCheck
    global v_XYZCheck
    global v_mtrlCheck
    global v_prog
    global selectedProgram
    global maxMenubarLen
    global version
    global onlineUpdateVer
    global updateFlag
    global menubar

    config_ini_path = "config.ini"
    if platform.system() == "Windows":
        config_ini_path = os.path.join(os.getenv("APPDATA"), "dend-mod-gui", "config.ini")

    getUpdateVer()

    root = tkinter.Tk()
    root.title(textSetting.textList["app"]["title"].format(version))
    root.option_add("*font", textSetting.textList["defaultFont"])
    root.geometry("1024x768")

    style = ttk.Style(root)
    style.configure(".", font=textSetting.textList["defaultFont"])

    menubar = tkinter.Menu(root)

    v_prog = tkinter.IntVar()

    progmenu = tkinter.Menu(menubar, tearoff=False)
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["SSUnity"], value=-1, variable=v_prog, command=lambda: callProgram("SSUnity"))
    progmenu.add_separator()
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["orgInfoEditor"], value=1, variable=v_prog, command=lambda: callProgram("orgInfoEditor"))
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["mdlBin"], value=2, variable=v_prog, command=lambda: callProgram("mdlBin"))
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["mdlinfo"], value=3, variable=v_prog, command=lambda: callProgram("mdlinfo"))
    progmenu.add_separator()
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["comicscript"], value=4, variable=v_prog, command=lambda: callProgram("comicscript"))
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["musicEditor"], value=5, variable=v_prog, command=lambda: callProgram("musicEditor"))
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["fvtMaker"], value=6, variable=v_prog, command=lambda: callProgram("fvtMaker"))
    progmenu.add_separator()
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["railEditor"], value=7, variable=v_prog, command=lambda: callProgram("railEditor"))
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["rsRail"], value=8, variable=v_prog, command=lambda: callProgram("rsRail"))
    progmenu.add_separator()
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["smf"], value=9, variable=v_prog, command=lambda: callProgram("smf"))
    progmenu.add_separator()
    progmenu.add_command(label=textSetting.textList["menu"]["appearance"]["rootFrame"], command=editRootFrameAppearance)
    progmenu.add_radiobutton(label=textSetting.textList["menu"]["program"]["exit"], value=-2, variable=v_prog, command=sys.exit)

    filemenu = tkinter.Menu(menubar, tearoff=False)
    filemenu.add_command(label=textSetting.textList["menu"]["file"]["loadFile"], command=loadFile)

    menubar.add_cascade(label=textSetting.textList["menu"]["program"]["name"], menu=progmenu)
    menubar.add_cascade(label=textSetting.textList["menu"]["file"]["name"], menu=filemenu)

    root.config(menu=menubar)

    if not os.path.exists(config_ini_path):
        writeDefaultConfig()

    readRootFrameAppearance()

    maxMenubarLen = menubar.index(tkinter.END)

    root.after(100, confirmUpdate)
    root.mainloop()


def execSaveRail(excelFile, railFile, quietFlag):
    if os.path.splitext(railFile)[1].lower() == ".den":
        return execSaveStageData(excelFile, railFile, quietFlag)

    import openpyxl
    import program.railEditor.dendDecrypt.RSdecrypt as dendRs
    import program.railEditor.dendDecrypt.CSdecrypt as dendCs
    import program.railEditor.dendDecrypt.BSdecrypt as dendBs
    import program.railEditor.dendDecrypt.LSdecrypt as dendLs
    from program.railEditor.importPy.excelWidget import ExcelWidget

    try:
        wb = openpyxl.load_workbook(excelFile, data_only=True)
        tabList = textSetting.textList["railEditor"]["railComboValue"]
        # ver
        ws = wb[tabList[0]]
        ver = ws.cell(1, 1).value

        oldVersionList = [
            "DEND_MAP_VER0100", "DEND_MAP_VER0101",
            "DEND_MAP_VER0102",
            "DEND_MAP_VER0110",
            "DEND_MAP_VER0300", "DEND_MAP_VER0400"
        ]

        if ver in oldVersionList:
            if ver == "DEND_MAP_VER0300" or ver == "DEND_MAP_VER0400":
                decryptFile = dendRs.RailDecrypt(railFile)
            elif ver == "DEND_MAP_VER0110":
                decryptFile = dendCs.RailDecrypt(railFile)
            elif ver == "DEND_MAP_VER0102":
                decryptFile = dendBs.RailDecrypt(railFile)
            elif ver == "DEND_MAP_VER0100" or ver == "DEND_MAP_VER0101":
                decryptFile = dendLs.RailDecrypt(railFile)
            excelWidget = ExcelWidget(decryptFile, None)

            if decryptFile.game == "LS":
                tabList = textSetting.textList["railEditor"]["railLsComboValue"]

            for tabName in tabList:
                if tabName not in wb.sheetnames:
                    errMsg = textSetting.textList["errorList"]["E95"].format(tabName)
                    mb.showerror(title=textSetting.textList["error"], message=errMsg)
                    return -3

            newByteArr = bytearray()

            if decryptFile.game == "LS":
                if not excelWidget.lsSave(wb, tabList, newByteArr):
                    if excelWidget.error == "":
                        errMsg = textSetting.textList["errorList"]["E14"]
                    else:
                        errMsg = excelWidget.error
                    mb.showerror(title=textSetting.textList["error"], message=errMsg)
                    return -4
            elif decryptFile.game == "BS":
                if not excelWidget.bsSave(wb, tabList, newByteArr):
                    if excelWidget.error == "":
                        errMsg = textSetting.textList["errorList"]["E14"]
                    else:
                        errMsg = excelWidget.error
                    mb.showerror(title=textSetting.textList["error"], message=errMsg)
                    return -4
            elif decryptFile.game == "CS":
                if not excelWidget.csSave(wb, tabList, newByteArr):
                    if excelWidget.error == "":
                        errMsg = textSetting.textList["errorList"]["E14"]
                    else:
                        errMsg = excelWidget.error
                    mb.showerror(title=textSetting.textList["error"], message=errMsg)
                    return -4
            elif decryptFile.game == "RS":
                if not excelWidget.rsSave(wb, tabList, newByteArr):
                    if excelWidget.error == "":
                        errMsg = textSetting.textList["errorList"]["E14"]
                    else:
                        errMsg = excelWidget.error
                    mb.showerror(title=textSetting.textList["error"], message=errMsg)
                    return -4
            w = open(decryptFile.filePath, "wb")
            w.write(newByteArr)
            w.close()
            if not quietFlag:
                mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I114"])
            return 0
    except Exception:
        errorLog(traceback.format_exc())
        mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
        return -1


def execSaveStageData(stageDataFile, denFile, quietFlag):
    from program.ssUnity.SSDecrypt.denDecrypt import DenDecrypt
    from program.ssUnity.ssUnity import loadExcelAndMerge
    decryptFile = DenDecrypt(denFile)
    if not decryptFile.open():
        decryptFile.printError()
        mb.showerror(title=textSetting.textList["error"], message=textSetting.textList["errorList"]["E14"])
        return -1
    if len(decryptFile.allList) == 1:
        dataName = decryptFile.allList[0][0]
        if dataName != "stagedata":
            errMsg = textSetting.textList["errorList"]["E105"] + denFile
            mb.showerror(title=textSetting.textList["error"], message=errMsg)
            return -5
        data = decryptFile.allList[0][-1]
        if os.path.splitext(stageDataFile)[1] == ".txt":
            with open(stageDataFile, "rb") as f:
                data.script = f.read()
        else:
            errMsgObj = {}
            if not loadExcelAndMerge(stageDataFile, data, errMsgObj):
                mb.showerror(title=textSetting.textList["error"], message=errMsgObj["message"])
                return -6
        data.save()
        with open(decryptFile.filePath, "wb") as w:
            w.write(decryptFile.env.file.save())
        if not quietFlag:
            mb.showinfo(title=textSetting.textList["success"], message=textSetting.textList["infoList"]["I51"])
        return 0
    else:
        errMsg = textSetting.textList["errorList"]["E105"] + denFile
        mb.showerror(title=textSetting.textList["error"], message=errMsg)
        return -6


if __name__ == "__main__":
    if len(sys.argv) >= 4:
        if sys.argv[1] in ["/saveRail", "/quietSaveRail", "/debugSaveRail"]:
            quietFlag = False
            debugFlag = False
            if sys.argv[1] == "/quietSaveRail":
                quietFlag = True
            elif sys.argv[1] == "/debugSaveRail":
                debugFlag = True
            excelFile = sys.argv[2]
            if not os.path.exists(excelFile):
                errMsg = textSetting.textList["errorList"]["E103"] + excelFile
                mb.showerror(title=textSetting.textList["error"], message=errMsg)
                sys.exit(-2)
            railFile = sys.argv[3]
            if os.path.splitext(railFile)[1].lower() == ".bin":
                if not os.path.exists(railFile):
                    w = open(railFile, "wb")
                    w.close()
            else:
                if not os.path.exists(railFile):
                    errMsg = textSetting.textList["errorList"]["E104"] + railFile
                    mb.showerror(title=textSetting.textList["error"], message=errMsg)
                    sys.exit(-2)
            if debugFlag:
                debugStr = "エクセルのパス\n{0}".format(os.path.abspath(excelFile))
                debugStr += "\n"
                debugStr += "ファイルのパス\n{0}".format(os.path.abspath(railFile))
                debugStr += "\nで確認しました。このまま実行しますか？"
                result = mb.askyesno(title=textSetting.textList["confirm"], message=debugStr)
                if not result:
                    sys.exit(0)
            sys.exit(execSaveRail(excelFile, railFile, quietFlag))
    guiMain()
